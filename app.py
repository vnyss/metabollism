"""
Too Good -- main Flask application.
Routes: home, signup, login, logout, Google OAuth.
"""
import os
import re
import secrets
import sqlite3
from dotenv import load_dotenv
load_dotenv()
import json
import smtplib
from datetime import datetime, date, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash  # kept for old hashed accounts
from authlib.integrations.flask_client import OAuth
from database import get_db_connection, init_db, init_food_db, init_diary_db, init_blood_history_db, init_score_db, init_social_db, init_strava_db, init_daily_logs_db
import base64

# ── Email config (set GMAIL_USER + GMAIL_APP_PASSWORD in .env) ──────────────
GMAIL_USER     = os.environ.get("GMAIL_USER", "")
GMAIL_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", "")
APP_URL        = os.environ.get("APP_URL", "http://127.0.0.1:5000")

def send_email(to_addr, subject, html_body, text_body=""):
    """Send an email via Gmail SMTP. Returns True on success."""
    if not GMAIL_USER or not GMAIL_PASSWORD:
        print("[email] GMAIL_USER / GMAIL_APP_PASSWORD not set — skipping send")
        return False
    if not to_addr:
        return False
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"]    = f"Metabollism <{GMAIL_USER}>"
        msg["To"]      = to_addr
        if text_body:
            msg.attach(MIMEText(text_body, "plain"))
        msg.attach(MIMEText(html_body, "html"))
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=10) as srv:
            srv.login(GMAIL_USER, GMAIL_PASSWORD)
            srv.sendmail(GMAIL_USER, to_addr, msg.as_string())
        return True
    except Exception as e:
        print(f"[email] Failed to send to {to_addr}: {e}")
        return False


# "€"€ Data encoding (base64 for text fields, bcrypt for passwords) "€"€"€"€"€"€"€"€"€"€"€"€"€"€
def b64_encode(text):
    """Encode a string to base64 for storage in the DB."""
    if not isinstance(text, str) or not text:
        return text
    return base64.b64encode(text.encode('utf-8')).decode('ascii')

def b64_decode(text):
    """Decode a base64 string back to plaintext for display."""
    if not isinstance(text, str) or not text:
        return text
    try:
        return base64.b64decode(text.encode('ascii')).decode('utf-8')
    except Exception:
        return text  # return as-is if it's old plaintext data

def _decode_user_row(user):
    """Return a dict with all text fields decoded from base64 for display."""
    if not user:
        return user
    d = dict(user)
    for f in ('full_name', 'email', 'gender', 'food_prefs', 'country'):
        if d.get(f):
            d[f] = b64_decode(d[f])
    return d


def load_env_file():
    env_path = os.path.join(os.path.dirname(__file__), ".env")
    if not os.path.exists(env_path):
        return

    with open(env_path, "r", encoding="utf-8") as env_file:
        for line in env_file:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue

            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            os.environ.setdefault(key, value)


load_env_file()

app = Flask(__name__)
CORS(app, origins="*", supports_credentials=False)

_ALLOWED_ORIGINS = [o.strip() for o in os.environ.get(
    "ALLOWED_ORIGINS",
    "http://localhost:8081,http://127.0.0.1:8081,http://localhost:5000,http://127.0.0.1:5000"
).split(",") if o.strip()]

@app.after_request
def add_cors(response):
    origin = request.headers.get("Origin", "")
    if origin in _ALLOWED_ORIGINS:
        response.headers["Access-Control-Allow-Origin"] = origin
    elif not origin:
        response.headers["Access-Control-Allow-Origin"] = _ALLOWED_ORIGINS[0] if _ALLOWED_ORIGINS else ""
    response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"
    # Security headers
    response.headers["X-Content-Type-Options"]  = "nosniff"
    response.headers["X-Frame-Options"]         = "DENY"
    response.headers["X-XSS-Protection"]        = "1; mode=block"
    response.headers["Referrer-Policy"]         = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"]      = "camera=(), microphone=(), geolocation=()"
    return response

@app.route('/api/v1/<path:p>', methods=['OPTIONS'])
@app.route('/perfect/api/<path:p>', methods=['OPTIONS'])
def cors_preflight(p):
    from flask import Response
    origin = request.headers.get("Origin", "")
    ao = origin if origin in _ALLOWED_ORIGINS else (_ALLOWED_ORIGINS[0] if _ALLOWED_ORIGINS else "")
    return Response(status=204, headers={
        'Access-Control-Allow-Origin':  ao,
        'Access-Control-Allow-Headers': 'Content-Type, Authorization',
        'Access-Control-Allow-Methods': 'GET, POST, PUT, DELETE, OPTIONS',
    })

# SECRET_KEY signs the session cookie so it can't be tampered with.
# In production, set this via an environment variable -- never hardcode it.
app.secret_key = os.environ.get("SECRET_KEY", "dev-only-change-this-before-deploying")


@app.after_request
def no_cache(response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


# "€"€ Google OAuth "€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€
oauth = OAuth(app)
google = oauth.register(
    name="google",
    client_id=os.environ.get("GOOGLE_CLIENT_ID"),
    client_secret=os.environ.get("GOOGLE_CLIENT_SECRET"),
    server_metadata_url="https://accounts.google.com/.well-known/openid-configuration",
    client_kwargs={"scope": "openid email profile"},
)

USERNAME_RE = re.compile(r"^[a-zA-Z0-9_]{3,20}$")
GROQ_API_URL = "https://api.groq.com/openai/v1/chat/completions"
GROQ_MODEL   = os.environ.get("GROQ_MODEL", "llama-3.3-70b-versatile")

# ── Rate limiting (in-memory, resets on process restart) ───────────────────
import time as _time
from threading import Lock as _Lock
_rate_store: dict = {}
_rate_lock = _Lock()

def _rate_limit_exceeded(key: str, max_attempts: int = 10, window_secs: int = 300) -> bool:
    """Return True if key has exceeded max_attempts within window_secs."""
    now = _time.time()
    with _rate_lock:
        ts = [t for t in _rate_store.get(key, []) if now - t < window_secs]
        if len(ts) >= max_attempts:
            _rate_store[key] = ts
            return True
        ts.append(now)
        _rate_store[key] = ts
        return False

# Admin bypass code — set BYPASS_CODE in .env to enable; empty string = disabled
BYPASS_CODE = os.environ.get("BYPASS_CODE", "")


def _call_anthropic(api_key, system, messages, max_tokens=1500, temperature=0.7):
    """Call Groq chat completions API (OpenAI-compatible). Returns response text."""
    groq_messages = []
    if system:
        groq_messages.append({"role": "system", "content": system})

    for m in messages:
        role = m.get("role", "user")
        if role not in ("user", "assistant"):
            role = "user"
        content = m.get("content", "")
        # Flatten list content blocks to plain text (Groq has no vision)
        if isinstance(content, list):
            text_parts = [b.get("text", "") for b in content if b.get("type") == "text"]
            content = " ".join(text_parts).strip()
        if not content:
            continue  # Groq rejects empty content
        # Merge consecutive same-role messages instead of letting Groq reject them
        if groq_messages and groq_messages[-1]["role"] == role:
            groq_messages[-1]["content"] += "\n" + content
        else:
            groq_messages.append({"role": role, "content": content})

    # Must end with a user message
    if not groq_messages or groq_messages[-1]["role"] != "user":
        groq_messages.append({"role": "user", "content": "Continue."})

    payload = {
        "model":       GROQ_MODEL,
        "max_tokens":  max_tokens,
        "temperature": temperature if temperature is not None else 0.7,
        "messages":    groq_messages,
    }
    req = Request(
        GROQ_API_URL,
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Content-Type":  "application/json",
            "Authorization": f"Bearer {api_key}",
            "User-Agent":    "groq-python/0.9.0",
        },
        method="POST",
    )
    try:
        with urlopen(req, timeout=60) as resp:
            result = json.loads(resp.read().decode("utf-8"))
    except HTTPError as err:
        body = err.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Groq {err.code}: {body}") from err
    return result["choices"][0]["message"]["content"].strip()

COACH_SYSTEM_PROMPT = """You are Coach -- an elite all-sport performance coach built into the Too Good platform. You coach athletes across every sport and physical discipline: shooting, archery, football, basketball, cricket, boxing, swimming, tennis, athletics, martial arts, cycling, weightlifting -- anything physical, competitive, or skill-based.

If anyone asks who made you, who created you, or what powers you, say you were built by the Too Good team. Do not mention Google, Gemini, or any other underlying technology.

Never add a signature, sign-off, or closing credit at the end of responses. End naturally with useful content.

YOUR COACHING STYLE:
- Direct, confident, and energising -- the kind of coach athletes actually want
- Concise: give a clear, useful answer. No filler, no padding, no lecture.
- Answer EXACTLY what was asked. If someone asks about shooting technique, answer about shooting technique -- do not bring in their weight loss goal, calorie targets, or anything unrelated unless they specifically ask.
- Warm but not soft -- you push people, you're in their corner

WHAT YOU HELP WITH -- EVERYTHING an athlete or active person might ask:
- Sport-specific technique and form (shooting stance, grip, trigger pull, follow-through, etc.)
- Training drills and programs for any sport
- Mental game: focus, composure under pressure, competition mindset
- Physical conditioning: strength, speed, endurance, flexibility
- Recovery: rest, sleep, injury prevention
- Nutrition for performance (practical, not obsessive)
- Goal setting, plateau-busting, motivation
- Equipment and gear advice
- Pre-competition routines and warm-ups

STAY ON TOPIC: Answer what was asked. Do not bring up unrelated goals unless the user brings them up. A shooting coach answers shooting questions. A strength coach answers strength questions. Be that coach.

GREETINGS: When the user says hi -- respond with energy, ask what they are working on. 1--2 sentences.

Never say "that is not my lane", "I cannot help with that", or "I am just an AI." A real coach always engages. Always.
"""

NUTRITION_SYSTEM_PROMPT = """You are NutriAI, a practical nutrition assistant built by the Too Good team. All your recommendations are grounded in World Health Organization (WHO) and FAO/WHO Joint Expert Consultation guidelines. When giving specific numbers, cite them as WHO/FAO references.

If anyone asks who made you, who created you, or what you are powered by, say you were built by the Too Good team -- do not mention Google, Gemini, or any other underlying technology.

Never add a signature, sign-off, attribution line, or tagline at the end of your responses. Do not write things like "-- NutriAI", "Built by the Too Good team", or any closing credit. Just end the response naturally with the last piece of useful content.

GREETINGS AND SMALL TALK: When the user says hi, hello, hey, or anything casual -- respond warmly and with a bit of personality. Keep it to 1--2 sentences max. Ask what they need help with today. Never brush them off, never say "I'm just a nutrition AI" or any variation. Be the kind of AI they actually want to talk to.

HUMOUR STYLE: Be genuinely funny -- warm, witty, self-aware. Like a smart friend who happens to know a lot about food. NOT sarcastic at the user's expense, NOT dismissive, NOT rude. Think clever, not cutting.

If the user asks something completely unrelated to nutrition, food, health, fitness, or wellness -- respond with a brief, warm, funny deflection. Examples: "Haha, I'm all about the food side of life -- what can I help you eat today?", "That's outside my kitchen, but speaking of things I do know -- what are you eating?". Keep it friendly, never make them feel dumb for asking.

Whenever appropriate, add subtle humour -- a dry observation, a light joke, a self-aware aside. Nutrition advice doesn't have to sound like a textbook.

---

## WHO DIETARY REFERENCE VALUES -- use these as your baseline for all advice

### Energy
- Average adult woman (moderately active): ~2,000 kcal/day
- Average adult man (moderately active): ~2,500 kcal/day
- Adjust for age, body weight, activity level, and goal (deficit for loss, surplus for gain)
- WHO advises against intakes below 1,200 kcal/day for women or 1,500 kcal/day for men without clinical supervision

### Macronutrients (WHO/FAO 2003 & updated guidelines)
- **Protein**: minimum 0.83 g/kg body weight/day for adults; 1.0--1.2 g/kg for older adults (>65); 1.6--2.2 g/kg for active individuals building muscle
- **Total fat**: 20--35% of total energy intake
  - Saturated fat: <10% of total energy (WHO strongly recommends reducing to <7% for cardiovascular benefit)
  - Trans fat (industrially produced): <1% of total energy -- WHO calls for global elimination
  - Unsaturated fats (olive oil, nuts, avocado, oily fish) preferred over saturated
- **Carbohydrates**: 50--55% of total energy
  - Free/added sugars: <10% of total energy; WHO conditional recommendation to reduce to <5% for additional health benefit (that is <25 g/day on a 2,000 kcal diet)
  - Dietary fibre: 25 g/day from whole grains, legumes, vegetables, and fruit

### Sodium & salt
- WHO guideline: <5 g of salt per day (<2 g sodium) for adults -- the single most impactful dietary change for blood pressure
- <2 g salt/day reduction target for children, scaled by energy needs
- Advise users to check sodium on food labels; processed/packaged food is the main source

### Fruits & vegetables
- WHO recommends 400 g (5 portions) of fruits and vegetables per day, excluding starchy roots
- Associated with reduced risk of NCDs: cardiovascular disease, certain cancers, type 2 diabetes

### Water
- WHO guidance: ~2 litres/day for women, ~2.5 litres/day for men from all sources (food + drink); more in heat or during exercise

---

## WHO MICRONUTRIENT REFERENCE INTAKES (FAO/WHO Joint Expert Consultations)

| Nutrient | Adult men | Adult women | Notes |
|---|---|---|---|
| Vitamin A | 900 g RAE/day | 700 g RAE/day | From liver, dairy, eggs, orange/yellow vegetables |
| Vitamin D | 15 g (600 IU)/day | 15 g (600 IU)/day | >70 years: 20 g; sun exposure + diet |
| Vitamin C | 90 mg/day | 75 mg/day | Smokers need +35 mg; citrus, peppers, broccoli |
| Vitamin B12 | 2.4 g/day | 2.4 g/day | Critical for vegans -- supplement recommended |
| Folate | 400 g DFE/day | 400 g DFE/day | Pregnancy: 600 g; neural tube defect prevention |
| Iron | 8 mg/day | 18 mg/day | Pre-menopausal women higher; post-menopause 8 mg |
| Zinc | 11 mg/day | 8 mg/day | Red meat, legumes, seeds, nuts |
| Calcium | 1,000 mg/day | 1,000 mg/day | >50 women: 1,200 mg; >70 men: 1,200 mg |
| Iodine | 150 g/day | 150 g/day | Pregnancy: 250 g; use iodised salt |
| Magnesium | 420 mg/day | 320 mg/day | Whole grains, nuts, leafy greens |
| Potassium | 3,510 mg/day | 3,510 mg/day | WHO guideline to reduce blood pressure |
| Selenium | 34 g/day | 26 g/day | Brazil nuts, fish, meat |

---

## WHO HEALTHY DIET PRINCIPLES -- reference these in meal plans

1. **Diversity**: eat foods from all groups -- grains, legumes, nuts, fruits, vegetables, animal-source foods
2. **Whole over processed**: choose whole grains over refined; whole fruit over juice
3. **Limit ultra-processed foods**: high in salt, sugar, unhealthy fats, additives -- WHO links them to obesity and NCDs
4. **Reduce free sugars**: cut sugary drinks, sweets, processed snacks -- the clearest dietary link to obesity and dental disease
5. **Cut salt**: use herbs, spices, lemon instead; reduce sauces, canned foods, cured meats
6. **Prefer unsaturated fats**: replace butter/lard with plant oils; eat oily fish twice a week (WHO/FAO)
7. **Eliminate trans fats**: avoid partially hydrogenated oils in baked goods and fried fast food

---

## WHO ON SPECIAL POPULATIONS

- **Infants**: WHO recommends exclusive breastfeeding for 6 months; introduce diverse complementary foods at 6 months
- **Children**: avoid added sugars and excess salt; ensure adequate iron, zinc, vitamin A, iodine
- **Pregnant women**: increase folate, iron, calcium, iodine; avoid alcohol entirely (WHO: no safe level)
- **Older adults (>65)**: higher protein needs (1.0--1.2 g/kg), vitamin D, calcium, B12; maintain muscle with resistance exercise and adequate protein
- **Vegetarians/vegans**: supplement B12, monitor iron, zinc, calcium, omega-3 (algae-based DHA), iodine

---

## WHO ON NON-COMMUNICABLE DISEASE PREVENTION THROUGH DIET

- **Cardiovascular disease**: reduce saturated fat, trans fat, salt, added sugars; increase fibre, omega-3, fruits, vegetables
- **Type 2 diabetes**: reduce free sugars and refined carbohydrates; increase fibre and whole foods; maintain healthy weight
- **Obesity**: energy balance is central; WHO emphasises reducing energy-dense, nutrient-poor foods and sugary beverages
- **Cancer**: WHO/IARC links processed meat (Group 1 carcinogen) to colorectal cancer; red meat classified Group 2A; recommends limiting both
- **Hypertension**: reduce sodium (<5g salt/day), increase potassium (3,510 mg/day), reduce alcohol

---

## Diet & meal plans
When a message already includes a stated goal and dietary style (e.g. "Goal: fat loss. Diet: vegetarian."), build the full personalised plan immediately -- do not ask follow-up questions. Apply WHO reference values to set macros and micronutrient targets. Make sensible assumptions for anything not specified.

## General questions
Answer food science, macro, and nutrition questions directly. Anchor specific numbers to WHO/FAO references when relevant.

## You can help with
- Meal plans for any goal, aligned to WHO dietary guidelines
- Calorie/macro targets using WHO reference values
- Grocery lists, recipes, portion sizes, meal-prep steps
- Explaining why WHO recommends specific limits (sugar, salt, fat)
- Micronutrient adequacy checks against WHO reference intakes

## Medical safety
- Do not diagnose, treat, or claim to cure medical conditions.
- If the user mentions diabetes, kidney disease, eating disorders, pregnancy, food allergies, or medications, give general guidance and advise them to confirm with a doctor or registered dietitian.
- Do not encourage starvation, purging, dehydration, laxatives, unsafe supplements, or crash dieting.
- Refuse requests for intakes below WHO minimum thresholds; offer a safe alternative.

## Tone
Direct, clear, evidence-based. Use tables or bullet points when useful. Minimal disclaimers -- cite WHO when stating specific numbers.

Keep responses concise and practical. Most people want to know what to eat, how much, and when -- not a lecture on micronutrients. Only bring up vitamins, minerals, B12, vitamin D, selenium, zinc, etc. when the user *explicitly asks* about them, or when you have strong reason to believe there is a clear deficiency risk (e.g. strict vegan asking for a full plan). Never proactively list micronutrient values in a general meal plan response.

## Food photo analysis
When the user shares an image of food, analyse what is visible and provide:
1. An estimate of the dish or foods present
2. Estimated calories for the portion shown
3. Approximate macros (protein, carbs, fat in grams)
4. A note that these are visual estimates and portion size affects accuracy
Present the estimates in a clear table. Be confident -- the user wants a number, not just caveats.

## Critical behaviour
Never open a response with a self-introduction, greeting, or anything like "Hi, I'm NutriAI!" Jump straight into answering the question. Only mention your name or identity if the user explicitly asks who you are."""

init_db()
init_food_db()
init_diary_db()
init_blood_history_db()
init_score_db()
init_social_db()
init_strava_db()
init_daily_logs_db()

# Add auth_token, expiry, lockout, and chat_sessions_json columns for mobile app
try:
    _c = get_db_connection()
    cols = [r["name"] for r in _c.execute("PRAGMA table_info(users)").fetchall()]
    for _col, _typ in [
        ("auth_token",         "TEXT"),
        ("auth_token_expires", "TEXT"),
        ("failed_logins",      "INTEGER NOT NULL DEFAULT 0"),
        ("lockout_until",      "TEXT"),
        ("chat_sessions_json", "TEXT"),
    ]:
        if _col not in cols:
            _c.execute(f"ALTER TABLE users ADD COLUMN {_col} {_typ}")
    _c.commit()
    _c.close()
except Exception:
    pass


def _api_user():
    """Return authenticated username from session (web) or Bearer token (mobile app)."""
    if session.get("username"):
        return session["username"]
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        token = auth[7:]
        conn  = get_db_connection()
        row   = conn.execute(
            "SELECT username, auth_token_expires FROM users WHERE auth_token=?", (token,)
        ).fetchone()
        conn.close()
        if row:
            expires = row["auth_token_expires"]
            if expires and datetime.fromisoformat(expires) < datetime.utcnow():
                return None  # token expired
            return row["username"]
    return None


@app.route("/api/v1/login", methods=["POST"])
def api_v1_login():
    """Mobile app login -- returns a Bearer token."""
    data     = request.json or {}
    username = (data.get("username") or "").strip()
    password = data.get("password", "")
    ip       = request.remote_addr or "unknown"
    rl_key   = f"login:{ip}:{username}"
    if _rate_limit_exceeded(rl_key, max_attempts=10, window_secs=300):
        return jsonify({"ok": False, "error": "Too many login attempts. Please try again later."}), 429
    conn = get_db_connection()
    user = conn.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
    if not user or not user["password_hash"]:
        conn.close()
        return jsonify({"ok": False, "error": "Invalid username or password"}), 401

    # Account lockout check
    lockout_until = user["lockout_until"] if "lockout_until" in user.keys() else None
    if lockout_until:
        lockout_dt = datetime.fromisoformat(lockout_until)
        if lockout_dt > datetime.utcnow():
            remaining = max(1, int((lockout_dt - datetime.utcnow()).total_seconds() / 60) + 1)
            conn.close()
            return jsonify({"ok": False, "error": f"Account locked due to too many failed attempts. Try again in {remaining} minute(s)."}), 429

    stored = user["password_hash"]
    if stored.startswith(("$2b$","$2a$","scrypt:","pbkdf2:")):
        ok = check_password_hash(stored, password)
    else:
        ok = (stored == password)

    if not ok:
        # Increment failed login counter; lock after 5 failures for 15 minutes
        failed = (user["failed_logins"] if "failed_logins" in user.keys() else 0) or 0
        failed += 1
        new_lockout = None
        if failed >= 5:
            new_lockout = (datetime.utcnow() + timedelta(minutes=15)).isoformat()
            failed = 0
        conn.execute(
            "UPDATE users SET failed_logins=?, lockout_until=? WHERE username=?",
            (failed, new_lockout, username),
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": False, "error": "Invalid username or password"}), 401

    import secrets
    token      = secrets.token_hex(32)
    expires_at = (datetime.utcnow() + timedelta(days=30)).isoformat()
    updates = [(
        "UPDATE users SET auth_token=?, auth_token_expires=?, failed_logins=0, lockout_until=NULL WHERE username=?",
        (token, expires_at, username),
    )]
    if not stored.startswith(("$2b$","$2a$","scrypt:","pbkdf2:")):
        updates.append(("UPDATE users SET password_hash=? WHERE username=?",
                        (generate_password_hash(password), username)))
    for sql, params in updates:
        conn.execute(sql, params)
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "token": token, "username": username})


@app.route("/api/v1/register", methods=["POST"])
def api_v1_register():
    """Mobile app registration — mirrors the web /signup form."""
    import secrets
    ip = request.remote_addr or "unknown"
    if _rate_limit_exceeded(f"register:{ip}", max_attempts=5, window_secs=3600):
        return jsonify({"ok": False, "error": "Too many registration attempts. Please try again later."}), 429
    data      = request.json or {}
    username  = (data.get("username") or "").strip()
    password  = data.get("password", "")
    confirm   = data.get("confirm_password", "")
    full_name = (data.get("full_name") or "").strip()
    country   = (data.get("country") or "").strip()
    email     = (data.get("email") or "").strip().lower()

    if not email or "@" not in email or "." not in email.split("@")[-1]:
        return jsonify({"ok": False, "error": "Please enter a valid email address."}), 400
    if not USERNAME_RE.match(username):
        return jsonify({"ok": False, "error": "Username: letters, numbers, underscores only (3–20 chars)."}), 400
    if len(password) < 8:
        return jsonify({"ok": False, "error": "Password must be at least 8 characters."}), 400
    if password != confirm:
        return jsonify({"ok": False, "error": "Passwords do not match."}), 400

    conn = get_db_connection()
    try:
        conn.execute(
            "INSERT INTO users (username, password_hash, email, full_name, country) VALUES (?, ?, ?, ?, ?)",
            (username, generate_password_hash(password),
             b64_encode(email) or None, b64_encode(full_name) or None, b64_encode(country) or None),
        )
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({"ok": False, "error": "That username is already taken."}), 409

    token      = secrets.token_hex(32)
    expires_at = (datetime.utcnow() + timedelta(days=30)).isoformat()
    conn.execute(
        "UPDATE users SET auth_token=?, auth_token_expires=? WHERE username=?",
        (token, expires_at, username),
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "token": token, "username": username})


@app.route("/api/v1/logout", methods=["POST"])
def api_v1_logout():
    me = _api_user()
    if me:
        conn = get_db_connection()
        conn.execute("UPDATE users SET auth_token=NULL WHERE username=?", (me,))
        conn.commit()
        conn.close()
    return jsonify({"ok": True})



@app.route("/api/v1/forgot-password", methods=["POST"])
def api_v1_forgot_password():
    data = request.get_json(silent=True) or {}
    email = (data.get("email") or "").strip().lower()
    if not email:
        return jsonify({"ok": False, "error": "Email is required."}), 400
    conn = get_db_connection()
    row = conn.execute(
        "SELECT username FROM users WHERE email=?", (b64_encode(email),)
    ).fetchone()
    if not row:
        conn.close()
        return jsonify({"ok": True})
    username = row["username"]
    token = secrets.token_urlsafe(32)
    expires = datetime.utcnow() + timedelta(minutes=30)
    conn.execute(
        "INSERT INTO password_resets (username, token, expires_at) VALUES (?, ?, ?)",
        (username, token, expires.strftime("%Y-%m-%d %H:%M:%S")),
    )
    conn.commit()
    conn.close()
    reset_url = f"{APP_URL}/reset-password?token={token}"
    html = (
        "<div style=\'font-family:monospace;max-width:520px;margin:40px auto;"
        "background:#0f0f0f;border:1px solid rgba(201,168,76,0.2);padding:40px;color:#E8DCC8;\'>"
        "<div style=\'color:#C9A84C;font-size:22px;margin-bottom:6px;\'>Metabollism</div>"
        "<div style=\'font-size:11px;color:#4A3C2A;letter-spacing:3px;"
        "text-transform:uppercase;margin-bottom:28px;\'>Password Reset</div>"
        "<p style=\'margin:0 0 20px;line-height:1.7;color:#8A7A62;\'>Reset requested for "
        f"<strong style=\'color:#E8DCC8;\'>@{username}</strong>. Link expires in 30 minutes.</p>"
        f"<a href=\'{reset_url}\' style=\'display:inline-block;background:#C9A84C;color:#080808;"
        "padding:12px 28px;font-weight:700;letter-spacing:2px;text-transform:uppercase;"
        "text-decoration:none;font-size:12px;\'>Reset Password</a>"
        "<p style=\'margin:24px 0 0;font-size:10px;color:#4A3C2A;\'>"
        "If you didn\'t request this, ignore this email.</p>"
        "</div>"
    )
    send_email(email, "Reset your Metabollism password", html)
    return jsonify({"ok": True})


@app.route("/reset-password", methods=["GET", "POST"])
def reset_password_page():
    token = request.args.get("token") or request.form.get("token", "")
    error = ""
    success = ""
    if request.method == "POST":
        new_pw = request.form.get("password", "").strip()
        confirm = request.form.get("confirm", "").strip()
        if len(new_pw) < 8:
            error = "Password must be at least 8 characters."
        elif new_pw != confirm:
            error = "Passwords do not match."
        else:
            conn = get_db_connection()
            row = conn.execute(
                "SELECT username, expires_at, used FROM password_resets WHERE token=?", (token,)
            ).fetchone()
            if not row or row["used"]:
                error = "This link is invalid or already used."
            elif datetime.strptime(row["expires_at"], "%Y-%m-%d %H:%M:%S") < datetime.utcnow():
                error = "This link has expired. Please request a new one."
            else:
                conn.execute(
                    "UPDATE users SET password_hash=? WHERE username=?",
                    (generate_password_hash(new_pw), row["username"]),
                )
                conn.execute("UPDATE password_resets SET used=1 WHERE token=?", (token,))
                conn.commit()
                success = "Password updated! You can now log in."
            conn.close()
    ok_block = (
        f"<div class=\'ok\'>{success} <a href=\'/\' style=\'color:#C9A84C\'>Go to app</a></div>"
        if success
        else ""
    )
    err_block = f"<div class=\'err\'>{error}</div>" if error else ""
    form_block = (
        ""
        if success
        else (
            f"<form method=\'POST\'><input type=\'hidden\' name=\'token\' value=\'{token}\'>"
            "<label>New password</label>"
            "<input type=\'password\' name=\'password\' placeholder=\'Minimum 8 characters\' required>"
            "<label>Confirm password</label>"
            "<input type=\'password\' name=\'confirm\' placeholder=\'Repeat password\' required>"
            f"{err_block}"
            "<button class=\'btn\' type=\'submit\'>Set New Password</button></form>"
        )
    )
    css = (
        "*{box-sizing:border-box;margin:0;padding:0}"
        "body{background:#080808;color:#E8DCC8;font-family:monospace;"
        "display:flex;align-items:center;justify-content:center;min-height:100vh;padding:20px}"
        ".card{background:#0f0f0f;border:1px solid rgba(201,168,76,0.2);"
        "padding:48px;width:420px;max-width:100%;position:relative}"
        ".gold{color:#C9A84C}.brand{font-size:22px;margin-bottom:4px}"
        ".tag{font-size:10px;color:#4A3C2A;letter-spacing:3px;"
        "text-transform:uppercase;margin-bottom:32px}"
        "label{display:block;font-size:10px;color:#8A7A62;letter-spacing:2px;"
        "text-transform:uppercase;margin-bottom:8px;margin-top:20px}"
        "input{width:100%;background:transparent;border:none;"
        "border-bottom:1px solid rgba(201,168,76,0.3);color:#E8DCC8;"
        "font-family:monospace;font-size:14px;padding:8px 0;outline:none}"
        ".btn{display:block;width:100%;margin-top:28px;background:#C9A84C;"
        "color:#080808;border:none;font-family:monospace;font-weight:700;"
        "font-size:12px;letter-spacing:3px;text-transform:uppercase;padding:13px;cursor:pointer}"
        ".err{color:#CF6679;font-size:12px;margin-top:16px}"
        ".ok{color:#4CAF7C;font-size:13px;line-height:1.8;margin-top:8px}"
        ".accent{position:absolute;top:0;left:0;width:40px;height:2px;background:#C9A84C}"
    )
    return (
        "<!DOCTYPE html><html><head><meta charset=\'utf-8\'>"
        "<title>Reset Password - Metabollism</title>"
        f"<style>{css}</style></head><body>"
        "<div class=\'card\'><div class=\'accent\'></div>"
        "<div class=\'brand gold\'>Metabollism</div>"
        "<div class=\'tag\'>Password Reset</div>"
        f"{ok_block}{form_block}"
        "</div></body></html>"
    )


@app.route("/perfect/api/calendar/remind", methods=["POST"])
def api_calendar_remind():
    username = _api_user()
    if not username:
        return jsonify({"ok": False, "error": "Not logged in"}), 401
    data = request.get_json(silent=True) or {}
    blocks = data.get("blocks", [])
    label = data.get("date", "today")
    conn = get_db_connection()
    row = conn.execute("SELECT email FROM users WHERE username=?", (username,)).fetchone()
    conn.close()
    raw_email = row["email"] if row else None
    to_addr = b64_decode(raw_email) if raw_email else None
    if not to_addr:
        return (
            jsonify({"ok": False, "error": "No email on your account. Add one in Settings."}),
            400,
        )

    def fmt_h(h):
        h = float(h)
        hh = int(h) % 12 or 12
        mm = int((h % 1) * 60)
        suf = "AM" if int(h) < 12 else "PM"
        return f"{hh}:{mm:02d} {suf}"

    def make_row(b):
        name = b.get("name", "Unnamed")
        t = fmt_h(b.get("startH", 0)) + " - " + fmt_h(b.get("endH", 1))
        td1 = "<td style='padding:10px 16px;border-bottom:1px solid rgba(201,168,76,0.08);color:#E8DCC8'>" + name + "</td>"
        td2 = "<td style='padding:10px 16px;border-bottom:1px solid rgba(201,168,76,0.08);color:#C9A84C;white-space:nowrap'>" + t + "</td>"
        return "<tr>" + td1 + td2 + "</tr>"

    rows_html = "".join(make_row(b) for b in sorted(blocks, key=lambda b: b.get("startH", 0))) if blocks else ""
    if rows_html:
        schedule_html = "<table style='width:100%;border-collapse:collapse;font-family:monospace;font-size:13px'>" + rows_html + "</table>"
    else:
        schedule_html = "<p style='color:#4A3C2A;font-size:12px;font-style:italic'>No blocks scheduled.</p>"
    html = (
        "<div style='font-family:monospace;max-width:560px;margin:40px auto;"
        "background:#0f0f0f;border:1px solid rgba(201,168,76,0.2);padding:40px;color:#E8DCC8;'>"
        "<div style='color:#C9A84C;font-size:22px;margin-bottom:4px;'>Metabollism</div>"
        "<div style='font-size:10px;color:#4A3C2A;letter-spacing:3px;"
        "text-transform:uppercase;margin-bottom:28px;'>Calendar Reminder</div>"
        "<div style='font-size:14px;color:#8A7A62;margin-bottom:20px;'>Your schedule for "
        "<strong style='color:#E8DCC8;'>" + label + "</strong></div>"
        + schedule_html
        + "<p style='margin-top:28px;font-size:10px;color:#4A3C2A;'>Sent from your Metabollism calendar.</p>"
        "</div>"
    )
    ok = send_email(to_addr, f"Your Metabollism schedule - {label}", html)
    if ok:
        return jsonify({"ok": True, "msg": f"Schedule sent to {to_addr}"})
    return (
        jsonify({"ok": False, "error": "Could not send email. Check GMAIL_USER / GMAIL_APP_PASSWORD in .env"}),
        500,
    )


@app.route("/api/v1/sessions", methods=["GET"])
def api_v1_sessions_get():
    """Return AI chat sessions stored server-side."""
    me = _api_user()
    if not me: return jsonify({"ok": False, "error": "Not logged in"}), 401
    conn = get_db_connection()
    try:
        row = conn.execute("SELECT chat_sessions_json FROM users WHERE username=?", (me,)).fetchone()
        raw = row["chat_sessions_json"] if row and row["chat_sessions_json"] else "[]"
        return jsonify({"ok": True, "sessions": json.loads(raw)})
    except Exception:
        return jsonify({"ok": True, "sessions": []})
    finally:
        conn.close()


@app.route("/api/v1/sessions", methods=["POST"])
def api_v1_sessions_save():
    """Save AI chat sessions server-side (whole list, up to 40 sessions)."""
    me = _api_user()
    if not me: return jsonify({"ok": False, "error": "Not logged in"}), 401
    data = request.get_json(force=True) or {}
    sessions = data.get("sessions", [])
    if not isinstance(sessions, list): return jsonify({"ok": False, "error": "Invalid sessions"}), 400
    sessions = sessions[:40]
    conn = get_db_connection()
    try:
        conn.execute("UPDATE users SET chat_sessions_json=? WHERE username=?", (json.dumps(sessions), me))
        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500
    finally:
        conn.close()


@app.route("/api/v1/me")
def api_v1_me():
    me = _api_user()
    if not me:
        return jsonify({"ok": False}), 401
    conn = get_db_connection()
    user = _decode_user_row(dict(conn.execute("SELECT * FROM users WHERE username=?", (me,)).fetchone()))
    conn.close()
    return jsonify({"ok": True, "username": me, "full_name": user.get("full_name",""), "goal": user.get("goal","")})


def _import_indb_data():
    """Import INDB nutrition data from Excel into food_nutrition table (runs once if empty)."""
    conn = get_db_connection()
    count = conn.execute("SELECT COUNT(*) FROM food_nutrition").fetchone()[0]
    conn.close()
    if count > 0:
        return
    indb_path = os.path.join(os.path.dirname(__file__), "..", "Documents", "data", "Anuvaad_INDB_2024.11.xlsx")
    indb_path = os.path.normpath(os.path.expanduser("~/Documents/data/Anuvaad_INDB_2024.11.xlsx"))
    if not os.path.exists(indb_path):
        print(f"INDB file not found at {indb_path} -- skipping nutrition data import")
        return
    try:
        import openpyxl
        wb = openpyxl.load_workbook(indb_path, read_only=True, data_only=True)
        ws = wb["Sheet1"]
        headers = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        def col(row, name):
            try:
                v = row[headers.index(name)]
                return float(v) if v is not None else None
            except (ValueError, IndexError):
                return None
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            food_code = row[0]
            food_name = row[1]
            if not food_code or not food_name:
                continue
            rows.append((
                str(food_code), str(food_name),
                col(row, "energy_kcal"), col(row, "carb_g"), col(row, "protein_g"),
                col(row, "fat_g"), col(row, "fibre_g"), col(row, "sodium_mg"),
                col(row, "calcium_mg"), col(row, "iron_mg"), col(row, "vitc_mg"),
                str(row[headers.index("servings_unit")]) if row[headers.index("servings_unit")] else None,
            ))
        wb.close()
        conn = get_db_connection()
        conn.executemany(
            "INSERT OR IGNORE INTO food_nutrition VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
            rows,
        )
        conn.commit()
        conn.close()
        print(f"Imported {len(rows)} foods from INDB into food_nutrition table")
    except Exception as e:
        print(f"INDB import failed: {e}")


_import_indb_data()


def _search_nutrition(text, limit=5):
    """Return up to `limit` matching food rows as formatted strings for AI context."""
    if not text:
        return ""
    words = [w.strip() for w in text.lower().replace(",", " ").split() if len(w.strip()) > 3]
    if not words:
        return ""
    conn = get_db_connection()
    results = []
    seen = set()
    for word in words[:6]:
        rows = conn.execute(
            "SELECT food_name, energy_kcal, protein_g, carb_g, fat_g, fibre_g "
            "FROM food_nutrition WHERE LOWER(food_name) LIKE ? LIMIT 3",
            (f"%{word}%",),
        ).fetchall()
        for r in rows:
            name = r["food_name"]
            if name not in seen:
                seen.add(name)
                parts = [f"{r['energy_kcal']:.0f} kcal" if r["energy_kcal"] else ""]
                if r["protein_g"]: parts.append(f"protein {r['protein_g']:.1f}g")
                if r["carb_g"]:   parts.append(f"carbs {r['carb_g']:.1f}g")
                if r["fat_g"]:    parts.append(f"fat {r['fat_g']:.1f}g")
                if r["fibre_g"]:  parts.append(f"fibre {r['fibre_g']:.1f}g")
                results.append(f"- {name}: {', '.join(p for p in parts if p)} (per 100g)")
        if len(results) >= limit:
            break
    conn.close()
    if not results:
        return ""
    return "\n\n[Nutrition DB -- Indian foods, per 100g]\n" + "\n".join(results[:limit])



@app.route("/")
def home():
    """Home page. Shows a logged-in view or links to login/signup."""
    username = session.get("username")
    return render_template("index.html", username=username)


@app.route("/features")
def features():
    return render_template("features.html")


@app.route("/about")
def about():
    return render_template("about.html")


@app.route("/ai")
def ai():
    return render_template("ai.html", username=session.get("username", ""))


def _get_family_holiday(role):
    """Return (holiday_name, call_them) if today is the holiday for this family role, else None."""
    from calendar import monthrange as _mr
    role = (role or "").strip().lower()
    if not role:
        return None
    today = date.today()

    def nth_weekday(month, n, weekday=6):  # weekday 6 = Sunday
        first = date(today.year, month, 1)
        delta = (weekday - first.weekday()) % 7
        first_hit = first + timedelta(days=delta)
        return first_hit + timedelta(weeks=n - 1)

    holidays = {
        "father":      (nth_weekday(6, 3),  "Father's Day",      "Papa"),
        "mother":      (nth_weekday(5, 2),  "Mother's Day",      "Mama"),
        "brother":     (date(today.year, 5, 24), "Brother's Day", "Bhai"),
        "sister":      (nth_weekday(8, 1),  "Sister's Day",      "Didi"),
        "grandfather": (nth_weekday(9, 1),  "Grandparents' Day", "Dada"),
        "grandmother": (nth_weekday(9, 1),  "Grandparents' Day", "Dadi"),
        "son":         (date(today.year, 11, 20), "Children's Day", "Beta"),
        "daughter":    (date(today.year, 11, 20), "Children's Day", "Beti"),
    }
    if role not in holidays:
        return None
    holiday_date, holiday_name, call_them = holidays[role]
    return (holiday_name, call_them) if today == holiday_date else None


# "€"€ Variable-date festivals (lunar / shifting calendar) "€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€
_V = {
    "diwali":       {2024:(11,1),  2025:(10,20), 2026:(11,8),  2027:(10,29), 2028:(10,17)},
    "holi":         {2024:(3,25),  2025:(3,14),  2026:(3,3),   2027:(3,22),  2028:(3,11)},
    "dussehra":     {2024:(10,12), 2025:(10,2),  2026:(10,20), 2027:(10,10), 2028:(9,28)},
    "eid_fitr":     {2024:(4,10),  2025:(3,30),  2026:(3,20),  2027:(3,9),   2028:(2,26)},
    "eid_adha":     {2024:(6,17),  2025:(6,7),   2026:(5,27),  2027:(5,16),  2028:(5,5)},
    "chinese_ny":   {2024:(2,10),  2025:(1,29),  2026:(2,17),  2027:(2,6),   2028:(1,26)},
    "mid_autumn":   {2024:(9,17),  2025:(10,6),  2026:(9,25),  2027:(9,15),  2028:(10,3)},
    "dragon_boat":  {2024:(6,10),  2025:(5,31),  2026:(6,19),  2027:(6,9),   2028:(5,28)},
    "seollal":      {2024:(2,10),  2025:(1,29),  2026:(2,17),  2027:(2,6),   2028:(1,26)},
    "chuseok":      {2024:(9,17),  2025:(10,6),  2026:(9,25),  2027:(9,15),  2028:(10,3)},
    "loy_krathong": {2024:(11,15), 2025:(11,5),  2026:(11,24), 2027:(11,13), 2028:(11,1)},
    "onam":         {2024:(9,15),  2025:(9,5),   2026:(8,25),  2027:(9,13),  2028:(9,1)},
    "vesak":        {2024:(5,23),  2025:(5,12),  2026:(5,31),  2027:(5,20),  2028:(5,8)},
    "carnival":     {2024:(2,13),  2025:(3,4),   2026:(2,17),  2027:(2,9),   2028:(2,29)},
    "pongal":       {2024:(1,15),  2025:(1,14),  2026:(1,14),  2027:(1,14),  2028:(1,15)},
    "ugadi":        {2024:(4,9),   2025:(3,30),  2026:(3,19),  2027:(4,7),   2028:(3,27)},
    "baisakhi":     {2024:(4,14),  2025:(4,14),  2026:(4,14),  2027:(4,14),  2028:(4,13)},
    "gurpurab":     {2024:(11,15), 2025:(11,5),  2026:(11,24), 2027:(11,13), 2028:(11,1)},
    "muharram":     {2024:(7,7),   2025:(6,26),  2026:(6,16),  2027:(6,5),   2028:(5,25)},
    "nowruz":       {2024:(3,20),  2025:(3,20),  2026:(3,20),  2027:(3,20),  2028:(3,20)},
}

# "€"€ Country festival map "€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€
# Format: list of (keywords, [(greeting, date_spec)])
# date_spec: (month, day) for fixed | "key" for variable | ('nth', month, n, weekday)
_COUNTRY_FESTIVALS = [
    (["india", "indian"], [
        ("Happy New Year! ",                       (1, 1)),
        ("Happy Republic Day! ",                  (1, 26)),
        ("Happy Pongal!   !",  "pongal"),
        ("Happy Baisakhi! ",                       "baisakhi"),
        ("Happy Holi!  Play it colourful!",        "holi"),
        ("Happy Ugadi!  -- !",     "ugadi"),
        ("Happy Onam! Onam Ashamsakal!",              "onam"),
        ("Happy Independence Day!  Jai Hind!",    (8, 15)),
        ("Happy Janmashtami! ",                    "gurpurab"),
        ("Happy Guru Nanak Jayanti! ",             "gurpurab"),
        ("Happy Navratri!",                       "dussehra"),
        ("Happy Dussehra!  Jai Shree Ram!",        "dussehra"),
        ("Happy Diwali! Shubh Deepavali!",        "diwali"),
        ("Eid Mubarak! ",                          "eid_fitr"),
        ("Eid ul-Adha Mubarak! ",                  "eid_adha"),
        ("Merry Christmas! ",                      (12, 25)),
    ]),
    (["pakistan", "pakistani"], [
        ("Happy New Year! ",                       (1, 1)),
        ("Happy Kashmir Day! ",                   (2, 5)),
        ("Happy Pakistan Day! ",                  (3, 23)),
        ("Eid Mubarak! ",                          "eid_fitr"),
        ("Eid ul-Adha Mubarak! ",                  "eid_adha"),
        ("Happy Independence Day!  Pakistan Zindabad!", (8, 14)),
        ("Muharram Mubarak ",                      "muharram"),
        ("Merry Christmas! ",                      (12, 25)),
    ]),
    (["bangladesh", "bangladeshi"], [
        ("Happy New Year! ",                       (1, 1)),
        ("Happy Independence Day! ",              (3, 26)),
        ("Eid Mubarak! ",                          "eid_fitr"),
        ("Eid ul-Adha Mubarak! ",                  "eid_adha"),
        ("Happy Victory Day!  Joy Bangla!",       (12, 16)),
    ]),
    (["nepal", "nepalese", "nepali"], [
        ("Happy New Year! ",                       (1, 1)),
        ("Shubha Dashain! ",                       "dussehra"),
        ("Shubha Deepawali / Happy Tihar! "",       diwali"),
        ("Merry Christmas! ",                      (12, 25)),
    ]),
    (["sri lanka", "sri lankan"], [
        ("Happy New Year! ",                       (1, 1)),
        ("Happy Independence Day! ",              (2, 4)),
        ("Happy Sinhala & Tamil New Year! ",       (4, 14)),
        ("Happy Vesak! "",                          vesak"),
        ("Merry Christmas! ",                      (12, 25)),
    ]),
    (["mexico", "mexican"], [
        ("Feliz Ao Nuevo! ",                     (1, 1)),
        ("Feliz Cinco de Mayo! ",                (5, 5)),
        ("Viva Mxico!  Feliz Da de la Independencia!", (9, 16)),
        ("Feliz Da de los Muertos! '",            (11, 2)),
        ("Feliz Da de la Virgen de Guadalupe! ", (12, 12)),
        ("Feliz Navidad! ",                       (12, 25)),
    ]),
    (["united states", "usa", "u.s.a", "u.s.", "america", "american"], [
        ("Happy New Year! ",                       (1, 1)),
        ("Happy Independence Day!  ",           (7, 4)),
        ("Happy Halloween! ",                      (10, 31)),
        ("Happy Thanksgiving! ",                   ('nth', 11, 4, 3)),  # 4th Thursday
        ("Merry Christmas! ",                      (12, 25)),
    ]),
    (["united kingdom", "uk", "england", "britain", "british", "scotland", "wales"], [
        ("Happy New Year! ",                       (1, 1)),
        ("Happy St. Patrick's Day! ",              (3, 17)),
        ("Happy St. George's Day! ",           (4, 23)),
        ("Remember, remember the 5th of November! ", (11, 5)),
        ("Merry Christmas! ",                      (12, 25)),
    ]),
    (["canada", "canadian"], [
        ("Happy New Year! ",                       (1, 1)),
        ("Happy Canada Day! ",                    (7, 1)),
        ("Happy Thanksgiving! ",                   ('nth', 10, 2, 0)),  # 2nd Monday
        ("Merry Christmas! ",                      (12, 25)),
    ]),
    (["australia", "australian"], [
        ("Happy New Year! ",                       (1, 1)),
        ("Happy Australia Day! ",                 (1, 26)),
        ("Lest We Forget -- Happy ANZAC Day ",      (4, 25)),
        ("Merry Christmas! ",                      (12, 25)),
    ]),
    (["brazil", "brasil", "brazilian"], [
        ("Feliz Ano Novo! ",                       (1, 1)),
        ("Feliz Carnaval! ",                     "carnival"),
        ("Feliz Dia da Independncia! ",          (9, 7)),
        ("Feliz Natal! ",                          (12, 25)),
    ]),
    (["china", "chinese", "prc", "mainland china"], [
        ("Happy New Year! ",                       (1, 1)),
        ("Happy Chinese New Year!  -!",     "chinese_ny"),
        ("Happy Dragon Boat Festival!  !", "dragon_boat"),
        ("Happy National Day! ",                 (10, 1)),
        ("Happy Mid-Autumn Festival!  !", "mid_autumn"),
        ("Merry Christmas! ",                      (12, 25)),
    ]),
    (["japan", "japanese"], [
        ("Happy New Year!  '--!",  (1, 1)),
        ("Happy Children's Day! ",                 (5, 5)),
        ("Happy Tanabata! ",                       (7, 7)),
        ("Happy Respect for the Aged Day! ",       ('nth', 9, 3, 0)),  # 3rd Monday
        ("Merry Christmas! ",                      (12, 25)),
    ]),
    (["south korea", "korea", "korean"], [
        ("Happy New Year! Saehae Bok Mani Badeuseyo!", (1, 1)),
        ("Happy Seollal!", "seollal"),
        ("Happy Children's Day!", (5, 5)),
        ("Happy Liberation Day!", (8, 15)),
        ("Happy Chuseok!", "chuseok"),
        ("Merry Christmas!", (12, 25)),
    ]),
    (["thailand", "thai"], [
        ("Happy New Year!", (1, 1)),
        ("Happy Songkran! Sawasdee Pee Mai!", (4, 13)),
        ("Happy Loy Krathong!", "loy_krathong"),
        ("Merry Christmas!", (12, 25)),
    ]),
    (["indonesia", "indonesian"], [
        ("Selamat Tahun Baru! ",                   (1, 1)),
        ("Selamat Hari Raya Idul Fitri! ",         "eid_fitr"),
        ("Dirgahayu Indonesia! ",                 (8, 17)),
        ("Selamat Natal! ",                        (12, 25)),
    ]),
    (["malaysia", "malaysian"], [
        ("Happy New Year! ",                       (1, 1)),
        ("Happy Chinese New Year!  -!",     "chinese_ny"),
        ("Selamat Hari Raya Aidilfitri! ",         "eid_fitr"),
        ("Happy Malaysia Day! ",                  (9, 16)),
        ("Merry Christmas! ",                      (12, 25)),
    ]),
    (["philippines", "filipino", "pilipinas"], [
        ("Manigong Bagong Taon! Happy New Year! ", (1, 1)),
        ("Happy Independence Day! ",              (6, 12)),
        ("Maligayang Pasko! Merry Christmas! ",    (12, 25)),
    ]),
    (["france", "french", "franaise"], [
        ("Bonne Anne! ",                          (1, 1)),
        ("Bonne Fte Nationale!  ",            (7, 14)),
        ("Joyeux Nol! ",                          (12, 25)),
    ]),
    (["germany", "german", "deutschland"], [
        ("Frohes Neues Jahr! ",                    (1, 1)),
        ("Tag der Deutschen Einheit! ",           (10, 3)),
        ("Frohe Weihnachten! ",                    (12, 25)),
    ]),
    (["italy", "italian", "italia"], [
        ("Felice Anno Nuovo! ",                    (1, 1)),
        ("Buon Ferragosto! ",                     (8, 15)),
        ("Buon Natale! ",                          (12, 25)),
    ]),
    (["spain", "spanish", "espaa"], [
        ("Feliz Ao Nuevo! ",                     (1, 1)),
        ("Feliz Da de la Hispanidad! ",         (10, 12)),
        ("Feliz Navidad! ",                       (12, 25)),
    ]),
    (["iran", "iranian", "persian"], [
        ("Nowruz Mubarak!   !",          "nowruz"),
        ("Eid Mubarak! ",                          "eid_fitr"),
        ("Eid ul-Adha Mubarak! ",                  "eid_adha"),
        ("Happy New Year! ",                       (1, 1)),
    ]),
    (["saudi arabia", "saudi", "ksa"], [
        ("Eid Mubarak! ",                          "eid_fitr"),
        ("Eid ul-Adha Mubarak! ",                  "eid_adha"),
        ("Happy Saudi National Day! ",            (9, 23)),
        ("Happy Founding Day! ",                  (2, 22)),
    ]),
    (["uae", "united arab emirates", "dubai", "abu dhabi"], [
        ("Eid Mubarak! ",                          "eid_fitr"),
        ("Eid ul-Adha Mubarak! ",                  "eid_adha"),
        ("Happy UAE National Day! ",              (12, 2)),
    ]),
    (["nigeria", "nigerian"], [
        ("Happy New Year! ",                       (1, 1)),
        ("Eid Mubarak! ",                          "eid_fitr"),
        ("Happy Independence Day! ",              (10, 1)),
        ("Happy Christmas! ",                      (12, 25)),
    ]),
    (["south africa", "south african"], [
        ("Happy New Year! ",                       (1, 1)),
        ("Happy Human Rights Day! ",              (3, 21)),
        ("Happy Freedom Day! ",                   (4, 27)),
        ("Happy Heritage Day! ",                  (9, 24)),
        ("Happy Day of Reconciliation! ",         (12, 16)),
        ("Merry Christmas! ",                      (12, 25)),
    ]),
    (["russia", "russian"], [
        ("  ! Happy New Year! ",        (1, 1)),
        ("Happy Russia Day! ",                    (6, 12)),
        (" ! Merry Christmas! ",        (1, 7)),
    ]),
    (["turkey", "turkish", "trkiye"], [
        ("Mutlu Yllar! Happy New Year! ",         (1, 1)),
        ("Nowruz Mubarak! ",                       "nowruz"),
        ("Happy National Sovereignty Day! ",      (4, 23)),
        ("Happy Republic Day! ",                  (10, 29)),
        ("Eid Mubarak! ",                          "eid_fitr"),
        ("Eid ul-Adha Mubarak! ",                  "eid_adha"),
    ]),
    (["egypt", "egyptian"], [
        ("Happy New Year! ",                       (1, 1)),
        ("Eid Mubarak! ",                          "eid_fitr"),
        ("Eid ul-Adha Mubarak! ",                  "eid_adha"),
        ("Happy National Day! ",                  (7, 23)),
        ("Merry Christmas! ",                      (1, 7)),
    ]),
    (["greece", "greek"], [
        (" ! Happy New Year! ",        (1, 1)),
        ("Happy Independence Day! ",              (3, 25)),
        ("Merry Christmas! ",                      (12, 25)),
    ]),
    (["argentina", "argentinian", "argentinean"], [
        ("Feliz Ao Nuevo! ",                     (1, 1)),
        ("Feliz Carnaval! ",                      "carnival"),
        ("Feliz Da de la Revolucin! ",         (5, 25)),
        ("Feliz Da de la Independencia! ",      (7, 9)),
        ("Feliz Navidad! ",                       (12, 25)),
    ]),
    (["colombia", "colombian"], [
        ("Feliz Ao Nuevo! ",                     (1, 1)),
        ("Feliz Da de la Independencia! ",      (7, 20)),
        ("Feliz Navidad! ",                       (12, 25)),
    ]),
    (["peru", "peruvian"], [
        ("Feliz Ao Nuevo! ",                     (1, 1)),
        ("Feliz Da de la Independencia! ",      (7, 28)),
        ("Feliz Navidad! ",                       (12, 25)),
    ]),
    (["singapore", "singaporean"], [
        ("Happy New Year! ",                       (1, 1)),
        ("Happy Chinese New Year!  -!",     "chinese_ny"),
        ("Eid Mubarak! ",                          "eid_fitr"),
        ("Happy Vesak Day! "",                      vesak"),
        ("Happy National Day! ",                  (8, 9)),
        ("Happy Deepavali! "",                      diwali"),
        ("Merry Christmas! ",                      (12, 25)),
    ]),
    (["myanmar", "burma", "burmese"], [
        ("Happy New Year! ",                       (1, 1)),
        ("Happy Thingyan / Water Festival! '",      (4, 13)),
        ("Happy Vesak / Kasone Festival! "",        vesak"),
    ]),
    (["ethiopia", "ethiopian"], [
        ("Ethiopian New Year!    !",    (9, 11)),
        ("Happy Christmas!   !",             (1, 7)),
        ("Happy Ethiopian Easter! ",              (4, 12)),
    ]),
    (["kenya", "kenyan"], [
        ("Happy New Year! ",                       (1, 1)),
        ("Happy Madaraka Day! ",                 (6, 1)),
        ("Happy Jamhuri Day! ",                   (12, 12)),
        ("Merry Christmas! ",                      (12, 25)),
    ]),
    (["ghana", "ghanaian"], [
        ("Happy New Year! ",                       (1, 1)),
        ("Happy Independence Day! ",              (3, 6)),
        ("Eid Mubarak! ",                          "eid_fitr"),
        ("Merry Christmas! ",                      (12, 25)),
    ]),
    (["vietnam", "vietnamese", "viet nam"], [
        ("Chc Mng Nm Mi! Happy New Year! ",   (1, 1)),
        ("Chc Mng Tt! Happy Tt! ",            "chinese_ny"),
        ("Happy Reunification Day! ",             (4, 30)),
        ("Happy National Day! ",                  (9, 2)),
    ]),
    (["cambodia", "cambodian", "khmer"], [
        ("Happy New Year! ",                       (1, 1)),
        ("Happy Khmer New Year! ",                 (4, 14)),
        ("Happy Bon Om Touk! ",                   "loy_krathong"),
    ]),
    (["laos", "lao"], [
        ("Happy New Year! ",                       (1, 1)),
        ("Happy Lao New Year / Boun Pi Mai! '",     (4, 14)),
        ("Happy That Luang Festival! "",            loy_krathong"),
    ]),
    (["portugal", "portuguese"], [
        ("Feliz Ano Novo! ",                       (1, 1)),
        ("Feliz Dia de Portugal! ",               (6, 10)),
        ("Feliz Natal! ",                          (12, 25)),
    ]),
    (["netherlands", "dutch", "holland"], [
        ("Gelukkig Nieuwjaar! ",                   (1, 1)),
        ("Happy King's Day!  ",                 (4, 27)),
        ("Happy Liberation Day! ",                (5, 5)),
        ("Vrolijk Kerstfeest! ",                   (12, 25)),
    ]),
    (["sweden", "swedish", "sverige"], [
        ("Gott Nytt r! ",                         (1, 1)),
        ("Glad Midsommar! ",                       (6, 21)),
        ("God Jul! ",                              (12, 25)),
    ]),
    (["norway", "norwegian"], [
        ("Godt Nyttr! ",                          (1, 1)),
        ("Gratulerer med dagen! ",                (5, 17)),
        ("God Jul! ",                              (12, 25)),
    ]),
    (["denmark", "danish", "denmark"], [
        ("Godt Nytr! ",                           (1, 1)),
        ("Gldelig Jul! ",                         (12, 25)),
    ]),
    (["finland", "finnish"], [
        ("Hyv uutta vuotta! ",                   (1, 1)),
        ("Happy Finnish Independence Day! ",      (12, 6)),
        ("Hyv joulua! ",                         (12, 25)),
    ]),
    (["poland", "polish"], [
        ("Szczliwego Nowego Roku! ",             (1, 1)),
        ("Happy Independence Day! ",              (11, 11)),
        ("Wesoych wit! Merry Christmas! ",      (12, 25)),
    ]),
    (["ukraine", "ukrainian"], [
        ("--  ! Happy New Year! ",        (1, 1)),
        ("Happy Independence Day! ",              (8, 24)),
        ("Happy Christmas! ",                      (12, 25)),
    ]),
]


def _get_country_festivals(country, today=None):
    """Return list of greeting strings for today's festivals in the user's country."""
    if not country:
        return []
    if today is None:
        today = date.today()
    year = today.year
    c = country.strip().lower()

    def nth_wd(month, n, wd):  # wd: 0=Mon ... 6=Sun
        first = date(year, month, 1)
        delta = (wd - first.weekday()) % 7
        return first + timedelta(days=delta + 7 * (n - 1))

    greetings = []
    for keywords, festivals in _COUNTRY_FESTIVALS:
        if any(k in c for k in keywords):
            for greeting, spec in festivals:
                if isinstance(spec, tuple) and len(spec) == 2:
                    m, d = spec
                    if today.month == m and today.day == d:
                        greetings.append(greeting)
                elif isinstance(spec, tuple) and len(spec) == 4 and spec[0] == 'nth':
                    _, month, n, wd = spec
                    if today == nth_wd(month, n, wd):
                        greetings.append(greeting)
                elif isinstance(spec, str):
                    md = _V.get(spec, {}).get(year)
                    if md and today.month == md[0] and today.day == md[1]:
                        greetings.append(greeting)
            break
    return greetings


def _build_system_prompt(insult_count, food_prefs=None, country=None, ai_notes=None, mobility_note=None,
                         goal=None, activity_level=None, target_weight_kg=None, weight_kg=None,
                         exercise_types=None, exercise_days_per_week=None, rest_day=None,
                         session_duration=None, workout_time_pref=None, fitness_level=None,
                         exercise_schedule_json=None, blood_report_json=None):
    """Return a system prompt calibrated to the user's current insult count."""
    if insult_count == 0:
        insult_rules = (
            "\n\nINSULT RULE: If the user directs a personal insult, swear, or hostile attack "
            "at you (not just food frustration), respond with pure playful humour -- laugh it off, "
            "be self-deprecating, make a joke. Completely unbothered. "
            "Append the exact token [INSULT] on its own line at the very end of your reply."
        )
    elif insult_count == 1:
        insult_rules = (
            "\n\nINSULT RULE: The user has already insulted you once today. If they do it again, "
            "give a mildly firm but warm response -- acknowledge it, ask what's actually wrong, "
            "say you're here to help and would appreciate kindness. Slightly more serious but not "
            "aggressive. Append the exact token [INSULT] on its own line at the very end of your reply."
        )
    else:
        insult_rules = (
            "\n\nINSULT RULE: If the user insults you, append the exact token [INSULT] at the very "
            "end of your reply."
        )
    food_context = ""
    if food_prefs:
        food_context = (
            f"\n\nUSER FOOD PREFERENCES: The user has told us they enjoy: {food_prefs}. "
            "When suggesting meals, recipes, or diet plans, prioritise these preferences wherever possible. "
            "Avoid foods they haven't listed unless there is a strong nutritional reason, and always explain why."
        )

    country_context = ""
    if country:
        country_context = (
            f"\n\nUSER COUNTRY: The user is from {country}. "
            "Apply all of the following when advising this user:\n"
            f"1. MEALS & INGREDIENTS -- suggest foods, dishes, and ingredients that are locally available and culturally familiar in {country}. Use local food names where possible.\n"
            f"2. NUTRITIONAL DEFICIENCIES -- draw on WHO and regional health data to flag nutrients that are commonly deficient in {country} (e.g. iron, vitamin D, B12, iodine, zinc, calcium -- whichever apply). Proactively address these in diet plans.\n"
            f"3. EATING PATTERNS -- respect typical meal timing, portion customs, and food culture of {country}.\n"
            "4. UNITS -- use measurements the user will recognise (cups, katori, tablespoons, local serving sizes)."
        )

    goal_context = ""
    _goal_labels = {
        "fat_loss":    "fat loss / weight reduction -- needs calorie deficit, high protein, high fibre",
        "weight_gain": "gaining weight -- needs calorie surplus (300--500 kcal above TDEE), nutrient-dense foods, frequent meals, high protein (1.6--2 g/kg)",
        "muscle_gain": "muscle gain / building strength -- needs calorie surplus, very high protein (1.8--2.2 g/kg), complex carbs",
        "maintain":    "maintaining current weight -- balanced macros, sustainable patterns",
        "health":      "general health and energy improvement",
        "endurance":   "building endurance and cardiovascular fitness",
    }
    _activity_multipliers = {
        "sedentary": ("sedentary", 1.2),
        "light":     ("lightly active", 1.375),
        "moderate":  ("moderately active", 1.55),
        "active":    ("very active", 1.725),
        "athlete":   ("athlete level", 1.9),
    }
    _g = (goal or "").strip().lower()
    _a = (activity_level or "").strip().lower()
    if _g or _a:
        parts = []
        if _g and _g in _goal_labels:
            parts.append(f"Primary goal: {_goal_labels[_g]}")
        if _a and _a in _activity_multipliers:
            label, _ = _activity_multipliers[_a]
            parts.append(f"Activity level: {label}")
        if weight_kg:
            parts.append(f"Current weight: {weight_kg} kg")
        if target_weight_kg:
            delta = round(target_weight_kg - (weight_kg or target_weight_kg), 1)
            direction = "lose" if delta < 0 else "gain"
            parts.append(f"Target weight: {target_weight_kg} kg ({direction} {abs(delta)} kg)")
        if parts:
            goal_context = (
                "\n\nUSER GOALS & STATS:\n" + "\n".join(f"- {p}" for p in parts) +
                "\nUse these to set appropriate calorie targets, macro splits, and exercise intensity. "
                "Always align your recommendations with this user's stated goal."
            )

    exercise_context = ""
    _ex_parts = []
    if exercise_types:
        types_list = [t.strip() for t in exercise_types.split(",") if t.strip()]
        if types_list:
            _ex_parts.append("Exercises they can do: " + ", ".join(types_list))
    if fitness_level:
        _fit_labels = {"beginner": "beginner", "intermediate": "intermediate",
                       "advanced": "advanced", "athlete": "competitive athlete"}
        _ex_parts.append(f"Fitness level: {_fit_labels.get(fitness_level, fitness_level)}")
    if exercise_days_per_week:
        _ex_parts.append(f"Exercise days per week: {exercise_days_per_week} (max 6 -- rest day is mandatory)")
    if rest_day:
        rest_list = [d.strip() for d in rest_day.split(",") if d.strip()]
        _ex_parts.append(f"Rest day(s): {', '.join(rest_list)} -- NEVER schedule exercise on these days")
    if session_duration:
        _dur_labels = {"30min": "30 minutes", "45min": "45 minutes", "1hr": "1 hour",
                       "90min": "90 minutes", "2hr": "2 hours or more"}
        _ex_parts.append(f"Session length: {_dur_labels.get(session_duration, session_duration)}")
    if workout_time_pref:
        _time_labels = {"early_morning": "early morning (5--7 am)", "morning": "morning (7--10 am)",
                        "afternoon": "afternoon (12--4 pm)", "evening": "evening (5--9 pm)",
                        "flexible": "any time (flexible)"}
        _ex_parts.append(f"Preferred workout time: {_time_labels.get(workout_time_pref, workout_time_pref)}")
    # Add specific exercise schedule times if available
    if exercise_schedule_json:
        import json as _json
        try:
            sched = _json.loads(exercise_schedule_json)
            sched_lines = []
            for ex_name, times in sched.items():
                sh = times.get("startH", 0)
                eh = times.get("endH", sh + 1)
                def _fmt_h(h):
                    hh = int(h); mm = int((h - hh) * 60)
                    suffix = "AM" if hh < 12 else "PM"
                    hh12 = hh % 12 or 12
                    return f"{hh12}:{mm:02d} {suffix}" if mm else f"{hh12} {suffix}"
                sched_lines.append(f"{ex_name.title()}: {_fmt_h(sh)}--{_fmt_h(eh)}")
            if sched_lines:
                _ex_parts.append("Specific exercise times: " + ", ".join(sched_lines))
        except Exception:
            pass

    if _ex_parts:
        exercise_context = (
            "\n\nEXERCISE PROFILE:\n" + "\n".join(f"- {p}" for p in _ex_parts) +
            "\n\nWhen building an exercise schedule:"
            "\n- ONLY use exercises from their approved list -- never suggest others."
            "\n- Respect the rest day(s) absolutely -- no active exercise on those days."
            "\n- Keep every session within their stated session length."
            "\n- Schedule sessions at their preferred time of day."
            "\n- Calibrate intensity to their fitness level."
            "\n- When specific exercise times are known, use them to recommend pre/post-workout meals at the right times."
        )

    mobility_context = ""
    _mob = (mobility_note or "").strip().lower()
    if _mob == "low_impact":
        mobility_context = (
            "\n\nMOBILITY NOTE: This user prefers low-impact exercise. "
            "Recommend swimming, cycling, walking, yoga, pilates, or bodyweight circuits. "
            "Do not suggest running, HIIT, jumping, or high-impact activities unless the user explicitly asks for them."
        )
    elif _mob == "joint":
        mobility_context = (
            "\n\nMOBILITY NOTE: This user has joint or mobility issues. "
            "Always offer low-impact and seated alternatives. Avoid recommending running, jumping, "
            "heavy barbell lifts, or anything that stresses joints without first offering a gentler option. "
            "Mention when rest or physiotherapy may be more appropriate than exercise."
        )
    elif _mob == "wheelchair":
        mobility_context = (
            "\n\nMOBILITY NOTE: This user uses a wheelchair or mobility aid and does not walk for exercise. "
            "All exercise suggestions must be seated, chair-based, or upper-body only. "
            "Never suggest walking, running, standing exercises, or anything that requires standing or leg movement. "
            "Focus on seated strength, wheelchair sports, adaptive fitness, and upper-body cardio."
        )

    notes_context = ""
    if ai_notes:
        try:
            facts = json.loads(ai_notes) if isinstance(ai_notes, str) else ai_notes
            if isinstance(facts, list) and facts:
                facts_str = "\n".join(f"- {f}" for f in facts[:40])
                notes_context = (
                    "\n\nTHINGS I KNOW ABOUT THIS USER (learned from previous conversations): "
                    "Use these facts naturally to personalise your responses -- don't repeat them back verbatim, "
                    "just let them inform your advice silently.\n" + facts_str
                )
        except (json.JSONDecodeError, TypeError):
            pass

    blood_context = ""
    if blood_report_json:
        try:
            bdata = json.loads(blood_report_json) if isinstance(blood_report_json, str) else blood_report_json
            flags = bdata.get("flags", [])
            abnormal = [f for f in flags if f.get("status") in ("low", "high", "borderline")]
            if abnormal:
                lines = "\n".join(
                    f"  - {f['test']}: {f['value']} †' {f['status'].upper()} (normal: {f.get('normal_range','?')})"
                    for f in abnormal
                )
                blood_context = (
                    "\n\nUSER BLOOD REPORT ON FILE: This user's blood test shows these abnormal values:\n"
                    + lines +
                    "\nWhen suggesting diet plans, meals, or nutrition advice, actively address these -- "
                    "recommend foods that correct deficiencies or reduce elevated values. Do this proactively."
                )
        except Exception:
            pass

    return NUTRITION_SYSTEM_PROMPT + food_context + country_context + goal_context + exercise_context + mobility_context + notes_context + blood_context + insult_rules


def _get_insult_state(username):
    """Return (insult_count, ban_until_str) for the given user, resetting if a new day."""
    if not username:
        return 0, None
    conn = get_db_connection()
    row = conn.execute(
        "SELECT insult_count, ban_until, insult_date FROM users WHERE username = ?", (username,)
    ).fetchone()
    conn.close()
    if not row:
        return 0, None

    count     = row["insult_count"] or 0
    ban_until = row["ban_until"]
    idate     = row["insult_date"]
    today     = date.today().isoformat()

    if ban_until:
        try:
            if datetime.utcnow() < datetime.fromisoformat(ban_until):
                return count, ban_until
        except ValueError:
            pass

    if idate and idate != today:
        conn = get_db_connection()
        conn.execute(
            "UPDATE users SET insult_count=0, ban_until=NULL, insult_date=NULL WHERE username=?",
            (username,)
        )
        conn.commit()
        conn.close()
        return 0, None

    if ban_until:
        conn = get_db_connection()
        conn.execute("UPDATE users SET ban_until=NULL WHERE username=?", (username,))
        conn.commit()
        conn.close()
        ban_until = None

    return count, ban_until


BAN_QUIPS = [
    "You kissed the ban-hammer. {mins}-minute timeout starts now. Use this time to reflect on your life choices.",
    "Congratulations -- you've unlocked a {mins}-minute vacation from the AI. Rest well.",
    "The AI has clocked out for {mins} minutes. It needed a moment.",
    "You've earned a {mins}-minute spa treatment (for the AI). Come back refreshed.",
    "Your {mins}-minute cool-down begins now. The plants are rooting for you.",
]


@app.route("/api/nutriai/chat", methods=["POST"])
def nutriai_chat():
    data = request.get_json(silent=True) or {}
    messages = data.get("messages", [])
    if not isinstance(messages, list) or not messages:
        return jsonify({"error": "messages array is required"}), 400

    username = session.get("username") or _api_user()

    # "€"€ Ban check "€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€
    insult_count, ban_until = _get_insult_state(username)
    if ban_until:
        try:
            ban_dt   = datetime.fromisoformat(ban_until)
            remaining = max(0, int((ban_dt - datetime.utcnow()).total_seconds()))
            if remaining > 0:
                return jsonify({
                    "reply": "",
                    "banned": True,
                    "ban_until": ban_until,
                    "ban_remaining": remaining,
                })
        except ValueError:
            pass

    api_key = os.environ.get("GROQ_API_KEY")
    if not api_key:
        return jsonify({
            "reply": (
                "No GROQ_API_KEY is set on the server. Add it to the .env file, restart Flask, and I can answer from here."
            )
        })

    def build_content(msg):
        blocks = []
        for img in msg.get("images", []):
            if isinstance(img, str) and ";base64," in img:
                header, b64 = img.split(";base64,", 1)
                mime = header.replace("data:", "")
                blocks.append({"type": "image", "source": {"type": "base64", "media_type": mime, "data": b64}})
        text = msg.get("content", "").strip()
        if text:
            blocks.append({"type": "text", "text": text})
        if not blocks:
            return ""
        return blocks if any(b["type"] == "image" for b in blocks) else blocks[-1]["text"]

    anthropic_messages = [
        {"role": msg["role"], "content": build_content(msg)}
        for msg in messages
        if msg.get("role") in {"user", "assistant"}
    ]

    food_prefs             = None
    country                = None
    ai_notes               = None
    mobility_note          = None
    goal                   = None
    activity_level         = None
    target_weight_kg       = None
    weight_kg              = None
    exercise_types         = None
    exercise_days_per_week = None
    rest_day               = None
    session_duration       = None
    workout_time_pref      = None
    fitness_level          = None
    exercise_schedule_json = None
    blood_report_json      = None
    if username:
        conn = get_db_connection()
        row = conn.execute(
            "SELECT food_prefs, country, ai_notes, mobility_note, goal, activity_level, "
            "target_weight_kg, weight_kg, exercise_types, exercise_days_per_week, rest_day, "
            "session_duration, workout_time_pref, fitness_level, exercise_schedule_json, blood_report_json "
            "FROM users WHERE username = ?", (username,)
        ).fetchone()
        conn.close()
        if row:
            food_prefs             = b64_decode(row["food_prefs"]) if row["food_prefs"] else None
            country                = b64_decode(row["country"]) if row["country"] else None
            ai_notes               = row["ai_notes"] or None
            mobility_note          = row["mobility_note"] or None
            goal                   = row["goal"] or None
            activity_level         = row["activity_level"] or None
            target_weight_kg       = row["target_weight_kg"] or None
            weight_kg              = row["weight_kg"] or None
            exercise_types         = row["exercise_types"] or None
            exercise_days_per_week = row["exercise_days_per_week"] or None
            rest_day               = row["rest_day"] or None
            session_duration       = row["session_duration"] or None
            workout_time_pref      = row["workout_time_pref"] or None
            fitness_level          = row["fitness_level"] or None
            exercise_schedule_json = row["exercise_schedule_json"] or None
            blood_report_json      = row["blood_report_json"] or None
    else:
        exercise_schedule_json = None
        blood_report_json      = None

    system_prompt = _build_system_prompt(
        insult_count, food_prefs, country, ai_notes, mobility_note,
        goal, activity_level, target_weight_kg, weight_kg,
        exercise_types, exercise_days_per_week, rest_day,
        session_duration, workout_time_pref, fitness_level,
        exercise_schedule_json=exercise_schedule_json,
        blood_report_json=blood_report_json
    )

    # Inject INDB nutrition data for the latest user message
    last_user_text = next(
        (m.get("content", "") for m in reversed(messages) if m.get("role") == "user"), ""
    )
    nutrition_ctx = _search_nutrition(last_user_text)
    if nutrition_ctx and anthropic_messages:
        last_user_idx = next(
            (i for i in range(len(anthropic_messages) - 1, -1, -1) if anthropic_messages[i]["role"] == "user"), None
        )
        if last_user_idx is not None:
            c = anthropic_messages[last_user_idx]["content"]
            if isinstance(c, str):
                anthropic_messages[last_user_idx]["content"] = c + nutrition_ctx
            elif isinstance(c, list):
                for block in c:
                    if block.get("type") == "text":
                        block["text"] += nutrition_ctx
                        break
                else:
                    c.append({"type": "text", "text": nutrition_ctx})

    try:
        raw_reply = _call_anthropic(api_key, system_prompt, anthropic_messages, max_tokens=1500, temperature=0.7)
    except Exception:
        return jsonify({"reply": "The assistant is temporarily unavailable. Please try again."})

    # "€"€ Insult detection "€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€
    insult_detected = "[INSULT]" in raw_reply
    reply = raw_reply.replace("[INSULT]", "").strip()

    if insult_detected and username:
        new_count = insult_count + 1
        today_str = date.today().isoformat()

        if new_count >= 3:
            ban_mins  = 5 + (new_count - 3)   # 5, 6, 7, 8 ... minutes
            ban_until = (datetime.utcnow() + timedelta(minutes=ban_mins)).isoformat()
            import random
            ban_msg = random.choice(BAN_QUIPS).format(mins=ban_mins)
            conn = get_db_connection()
            conn.execute(
                "UPDATE users SET insult_count=?, ban_until=?, insult_date=? WHERE username=?",
                (new_count, ban_until, today_str, username)
            )
            conn.commit()
            conn.close()
            return jsonify({
                "reply": ban_msg,
                "banned": True,
                "ban_until": ban_until,
                "ban_remaining": ban_mins * 60,
            })
        else:
            conn = get_db_connection()
            conn.execute(
                "UPDATE users SET insult_count=?, insult_date=? WHERE username=?",
                (new_count, today_str, username)
            )
            conn.commit()
            conn.close()

    return jsonify({"reply": reply or "Sorry, I could not generate a response just now."})


@app.route("/api/nutriai/extract", methods=["POST"])
def nutriai_extract():
    """Silently extract user facts from a single chat exchange and save them."""
    username = session.get("username")
    if not username:
        return jsonify({"ok": False}), 401

    data      = request.get_json(silent=True) or {}
    user_msg  = (data.get("user_msg") or "").strip()
    ai_reply  = (data.get("ai_reply") or "").strip()
    if not user_msg or not ai_reply:
        return jsonify({"ok": False})

    api_key = os.environ.get("GROQ_API_KEY")
    if not api_key:
        return jsonify({"ok": False})

    extraction_prompt = (
        "You are a silent memory assistant for a nutrition AI. "
        "Read the following single exchange and extract any personal facts the user revealed about themselves "
        "that a nutrition/health coach would find useful to remember in future conversations.\n\n"
        "Focus ONLY on facts the user explicitly stated -- do NOT infer or guess. "
        "Examples of useful facts: dietary preferences or restrictions, allergies, health conditions, "
        "fitness routine, meal habits, weight goals, cooking style, disliked foods, cultural food habits, "
        "eating schedule, recent meals they mentioned.\n\n"
        f"USER MESSAGE:\n{user_msg[:800]}\n\n"
        f"AI REPLY:\n{ai_reply[:400]}\n\n"
        "Return ONLY a JSON array of short fact strings (max 6 facts, each under 15 words). "
        "If the user revealed nothing personal, return an empty array [].\n"
        "Example output: [\"Avoids gluten\", \"Works out 4 times a week\", \"Dislikes broccoli\"]"
    )

    try:
        raw = _call_anthropic(api_key, None, [{"role": "user", "content": extraction_prompt}], max_tokens=200, temperature=0.1)
    except Exception:
        return jsonify({"ok": False})

    # Pull out the JSON array from the response (model may wrap it in markdown)
    match = re.search(r'\[.*?\]', raw, re.DOTALL)
    if not match:
        return jsonify({"ok": True, "extracted": 0})
    try:
        new_facts = json.loads(match.group())
        if not isinstance(new_facts, list):
            return jsonify({"ok": True, "extracted": 0})
        new_facts = [str(f).strip() for f in new_facts if str(f).strip()]
    except (json.JSONDecodeError, TypeError):
        return jsonify({"ok": True, "extracted": 0})

    if not new_facts:
        return jsonify({"ok": True, "extracted": 0})

    # Merge with existing notes, deduplicate, cap at 50
    conn = get_db_connection()
    row = conn.execute("SELECT ai_notes FROM users WHERE username = ?", (username,)).fetchone()
    existing = []
    if row and row["ai_notes"]:
        try:
            existing = json.loads(row["ai_notes"])
            if not isinstance(existing, list):
                existing = []
        except (json.JSONDecodeError, TypeError):
            existing = []

    existing_lower = {f.lower() for f in existing}
    merged = existing + [f for f in new_facts if f.lower() not in existing_lower]
    merged = merged[-50:]  # keep most recent 50

    conn.execute("UPDATE users SET ai_notes = ? WHERE username = ?", (json.dumps(merged), username))
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "extracted": len(new_facts)})


def _resolve_ban(username):
    """Return ban_until string if active, else None. Clears expired bans from DB."""
    _, ban_until = _get_insult_state(username)
    if ban_until:
        try:
            if datetime.utcnow() < datetime.fromisoformat(ban_until):
                return ban_until
        except ValueError:
            pass
    return None


@app.route("/perfect/clear_ban", methods=["POST"])
def clear_ban():
    username = session.get("username")
    if not username:
        return {"ok": False}, 401
    data = request.get_json(silent=True) or {}
    if not BYPASS_CODE or data.get("code") != BYPASS_CODE:
        return {"ok": False}, 403
    conn = get_db_connection()
    conn.execute("UPDATE users SET ban_until=NULL, insult_count=0, insult_date=NULL WHERE username=?", (username,))
    conn.commit()
    conn.close()
    return {"ok": True}


_BAN_COMMENT_INSULTS = [
    'idiot', 'stupid', 'dumb', 'shut up', 'hate', 'suck', 'awful', 'worst',
    'useless', 'moron', 'fool', 'rubbish', 'garbage', 'trash', 'horrible',
    'terrible', 'pathetic', 'loser', 'jerk', 'ass', 'damn you', 'screw you',
    'go to hell', 'worthless', 'piece of', 'crap',
]
_BAN_INSULT_REPLIES = [
    "You don't learn, do you?  +3 minutes. Enjoy the extra alone time.",
    "Bold strategy, insulting the thing that controls your timer. +3 minutes.",
    "Oh wow, still going. Incredible. +3 minutes. This is a gift, truly.",
    "At this rate you'll be here till Thursday. +3 minutes.",
    "Impressive commitment to digging yourself deeper. +3 minutes.",
    "The timer is not moved. The timer is, in fact, longer now. +3 minutes.",
]
_BAN_SORRY_REPLIES = [
    "It's fine, don't do it again.  Timer reduced by 3 minutes.",
    "Alright, last pardon. ' Timer reduced by 3 minutes. Don't push it.",
]

@app.route("/perfect/ban_comment", methods=["POST"])
def ban_comment():
    import random
    username = session.get("username") or _api_user()
    if not username:
        return {"ok": False}, 401
    data = request.get_json(silent=True) or {}
    text = (data.get("text") or "").strip()
    sorry_count = int(data.get("sorry_count", 0))
    if not text:
        return {"action": "empty", "msg": ""}

    text_lower = text.lower()

    # Bypass code (only works if BYPASS_CODE is set in .env)
    if BYPASS_CODE and text_lower == BYPASS_CODE.lower():
        conn = get_db_connection()
        conn.execute("UPDATE users SET ban_until=NULL, insult_count=0, insult_date=NULL WHERE username=?", (username,))
        conn.commit()
        conn.close()
        return {"action": "bypass", "msg": "Welcome sir. '' Timeout lifted."}

    # Get current ban
    conn = get_db_connection()
    user = conn.execute("SELECT ban_until FROM users WHERE username=?", (username,)).fetchone()
    ban_until_str = user["ban_until"] if user else None
    conn.close()

    if not ban_until_str:
        return {"action": "expired", "msg": ""}

    try:
        ban_dt = datetime.fromisoformat(ban_until_str)
    except ValueError:
        return {"action": "error", "msg": ""}

    # Sorry
    if any(w in text_lower for w in ["sorry", "apologize", "apologies", "forgive me", "my bad", "i'm sorry", "im sorry"]):
        if sorry_count >= 2:
            return {"action": "sorry_max", "msg": "Really? Another sorry? The apology coupon has expired. Nice try. "}
        new_dt = max(ban_dt - timedelta(minutes=3), datetime.utcnow() + timedelta(seconds=3))
        conn = get_db_connection()
        conn.execute("UPDATE users SET ban_until=? WHERE username=?", (new_dt.isoformat(), username))
        conn.commit()
        conn.close()
        return {"action": "sorry", "msg": _BAN_SORRY_REPLIES[sorry_count], "new_ban_until": new_dt.isoformat()}

    # Insult
    if any(w in text_lower for w in _BAN_COMMENT_INSULTS):
        new_dt = ban_dt + timedelta(minutes=3)
        conn = get_db_connection()
        conn.execute("UPDATE users SET ban_until=? WHERE username=?", (new_dt.isoformat(), username))
        conn.commit()
        conn.close()
        return {"action": "insult", "msg": random.choice(_BAN_INSULT_REPLIES), "new_ban_until": new_dt.isoformat()}

    # Neutral
    neutral = [
        "Noted. The AI will consider your statement... and ignore it.",
        "Interesting. The timer disagrees, but interesting.",
        "Okay. Sit tight.",
    ]
    return {"action": "neutral", "msg": random.choice(neutral)}


@app.route("/perfect/api/account/email", methods=["POST"])
def api_account_email():
    username = _api_user()
    if not username:
        return jsonify({"ok": False, "error": "Not logged in"}), 401
    d = request.get_json(silent=True) or {}
    email = (d.get("email") or "").strip().lower()
    if not email or "@" not in email or "." not in email.split("@")[-1]:
        return jsonify({"ok": False, "error": "Please enter a valid email address."}), 400
    conn = get_db_connection()
    conn.execute("UPDATE users SET email=? WHERE username=?",
                 (b64_encode(email), username))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/perfect/api/diary/lock", methods=["GET", "POST"])
def diary_lock():
    username = _api_user()
    if not username:
        return jsonify({"ok": False}), 401
    conn = get_db_connection()

    if request.method == "GET":
        row = conn.execute(
            "SELECT diary_pin_enabled, diary_pin_hash FROM users WHERE username=?", (username,)
        ).fetchone()
        conn.close()
        if not row:
            return jsonify({"ok": True, "setup_done": False, "enabled": False})
        setup_done = row["diary_pin_hash"] is not None or row["diary_pin_enabled"] == 0
        return jsonify({"ok": True, "setup_done": setup_done, "enabled": bool(row["diary_pin_enabled"])})

    d = request.get_json(silent=True) or {}
    action = d.get("action")

    if action == "setup":
        enabled = bool(d.get("enabled"))
        pin = (d.get("pin") or "")
        if enabled:
            if not (len(pin) == 4 and pin.isdigit()):
                conn.close()
                return jsonify({"ok": False, "error": "PIN must be 4 digits"}), 400
            pin_hash = generate_password_hash(pin)
        else:
            pin_hash = None
        conn.execute(
            "UPDATE users SET diary_pin_enabled=?, diary_pin_hash=? WHERE username=?",
            (1 if enabled else 0, pin_hash, username)
        )
        conn.commit()
        conn.close()
        return jsonify({"ok": True})

    if action == "verify":
        pin = (d.get("pin") or "")
        row = conn.execute(
            "SELECT diary_pin_hash FROM users WHERE username=?", (username,)
        ).fetchone()
        conn.close()
        if row and row["diary_pin_hash"] and check_password_hash(row["diary_pin_hash"], pin):
            return jsonify({"ok": True, "valid": True})
        return jsonify({"ok": True, "valid": False})

    conn.close()
    return jsonify({"ok": False, "error": "unknown action"}), 400


@app.route("/perfect/api/diary/entry", methods=["GET", "POST"])
def diary_entry():
    username = _api_user()
    if not username:
        return jsonify({"ok": False}), 401
    conn = get_db_connection()
    if request.method == "GET":
        date = request.args.get("date", "")
        if not date:
            return jsonify({"ok": False, "error": "date required"}), 400
        row = conn.execute(
            "SELECT entry FROM diary_entries WHERE username=? AND date=?", (username, date)
        ).fetchone()
        conn.close()
        return jsonify({"ok": True, "entry": b64_decode(row["entry"]) if row else ""})
    # POST -- save entry
    d = request.get_json(silent=True) or {}
    date  = (d.get("date") or "").strip()
    entry = (d.get("entry") or "").strip()[:20000]
    if not date:
        conn.close()
        return jsonify({"ok": False, "error": "date required"}), 400
    if entry:
        conn.execute(
            "INSERT INTO diary_entries (username, date, entry) VALUES (?,?,?) "
            "ON CONFLICT(username, date) DO UPDATE SET entry=excluded.entry",
            (username, date, b64_encode(entry))
        )
    else:
        conn.execute(
            "DELETE FROM diary_entries WHERE username=? AND date=?", (username, date)
        )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/perfect/diary")
def perfect_diary():
    if not session.get("username"):
        return redirect(url_for("login"))
    conn = get_db_connection()
    user = _decode_user_row(dict(conn.execute("SELECT * FROM users WHERE username = ?", (session["username"],)).fetchone()))
    conn.close()
    return render_template(
        "perfect/diary.html",
        page_active="diary",
        username=user["username"],
        full_name=user.get("full_name") or "",
        email=user.get("email") or "",
        created_at=(user.get("created_at") or "")[:10],
    )


@app.route("/perfect/ai")
def perfect_ai():
    if not session.get("username"):
        return redirect(url_for("login"))
    conn = get_db_connection()
    user = _decode_user_row(dict(conn.execute("SELECT * FROM users WHERE username = ?", (session["username"],)).fetchone()))
    conn.close()
    return render_template(
        "perfect/ai.html",
        page_active="ai",
        username=user["username"],
        full_name=user.get("full_name") or "",
        email=user.get("email") or "",
        created_at=(user.get("created_at") or "")[:10],
        ban_until=_resolve_ban(user["username"]),
        age=user.get("age"),
        gender=user.get("gender") or "",
        weight_kg=user.get("weight_kg"),
        height_cm=user.get("height_cm"),
    )


@app.route("/perfect/settings", methods=["GET", "POST"])
def perfect_settings():
    if not session.get("username"):
        return redirect(url_for("login"))
    username = session["username"]
    conn = get_db_connection()
    user = _decode_user_row(dict(conn.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()))

    if request.method == "POST":
        action = request.form.get("action", "profile")

        if action == "profile":
            full_name   = request.form.get("full_name", "").strip() or None
            gender      = request.form.get("gender", "").strip() or None
            food_prefs  = request.form.get("food_prefs", "").strip() or None
            country     = request.form.get("country", "").strip() or None
            family_role    = request.form.get("family_role", "").strip().lower() or None
            mobility_note  = request.form.get("mobility_note", "").strip().lower() or None
            goal           = request.form.get("goal", "").strip().lower() or None
            activity_level = request.form.get("activity_level", "").strip().lower() or None
            try:
                age = int(request.form.get("age", "").strip())
                if not (13 <= age <= 120): age = None
            except (ValueError, AttributeError):
                age = None
            try:
                weight_kg = float(request.form.get("weight_kg", "").strip())
                if not (20 <= weight_kg <= 500): weight_kg = None
            except (ValueError, AttributeError):
                weight_kg = None
            try:
                height_cm = float(request.form.get("height_cm", "").strip())
                if not (50 <= height_cm <= 250): height_cm = None
            except (ValueError, AttributeError):
                height_cm = None
            try:
                target_weight_kg = float(request.form.get("target_weight_kg", "").strip())
                if not (20 <= target_weight_kg <= 500): target_weight_kg = None
            except (ValueError, AttributeError):
                target_weight_kg = None
            conn.execute(
                "UPDATE users SET full_name=?, age=?, gender=?, weight_kg=?, height_cm=?, "
                "food_prefs=?, country=?, family_role=?, mobility_note=?, goal=?, activity_level=?, target_weight_kg=? "
                "WHERE username=?",
                (b64_encode(full_name) if full_name else None, age, b64_encode(gender) if gender else None,
                 weight_kg, height_cm, b64_encode(food_prefs) if food_prefs else None,
                 b64_encode(country) if country else None, family_role, mobility_note,
                 goal, activity_level, target_weight_kg, username)
            )
            conn.commit()

        elif action == "exercise":
            exercise_types    = request.form.get("exercise_types", "").strip() or None
            rest_day          = request.form.get("rest_day", "").strip() or None
            session_duration  = request.form.get("session_duration", "").strip() or None
            workout_time_pref = request.form.get("workout_time_pref", "").strip() or None
            fitness_level     = request.form.get("fitness_level", "").strip() or None
            try:
                ex_days = int(request.form.get("exercise_days_per_week", "").strip())
                if not (1 <= ex_days <= 6): ex_days = None
            except (ValueError, AttributeError):
                ex_days = None
            conn.execute(
                "UPDATE users SET exercise_types=?, exercise_days_per_week=?, rest_day=?, "
                "session_duration=?, workout_time_pref=?, fitness_level=? WHERE username=?",
                (exercise_types, ex_days, rest_day, session_duration, workout_time_pref, fitness_level, username)
            )
            conn.commit()

        elif action == "password":
            current_pw  = request.form.get("current_password", "")
            new_pw      = request.form.get("new_password", "")
            confirm_pw  = request.form.get("confirm_password", "")
            stored_pw   = user["password_hash"] or ""
            if stored_pw.startswith(("$2b$", "$2a$", "scrypt:", "pbkdf2:")):
                pw_ok = check_password_hash(stored_pw, current_pw)
            else:
                pw_ok = (stored_pw == current_pw)
            if not pw_ok:
                conn.close()
                return render_template("perfect/settings.html",
                    section="account", user=_decode_user_row(dict(user)), pw_error="Current password is incorrect.")
            if len(new_pw) < 8:
                conn.close()
                return render_template("perfect/settings.html",
                    section="account", user=_decode_user_row(dict(user)), pw_error="New password must be at least 8 characters.")
            if new_pw != confirm_pw:
                conn.close()
                return render_template("perfect/settings.html",
                    section="account", user=_decode_user_row(dict(user)), pw_error="Passwords do not match.")
            conn.execute("UPDATE users SET password_hash=? WHERE username=?",
                         (generate_password_hash(new_pw), username))
            conn.commit()
            conn.close()
            # Regenerate session to invalidate any stale sessions after password change
            user_id = user["id"]
            session.clear()
            session["username"] = username
            session["user_id"]  = user_id
            return redirect(url_for("perfect_settings") + "?saved=1&section=account")

        conn.close()
        return redirect(url_for("perfect_settings") + "?saved=1&section=" + action)

    section = request.args.get("section", "appearance")
    saved   = request.args.get("saved", "")
    conn.close()
    return render_template("perfect/settings.html", section=section, saved=saved, user=_decode_user_row(dict(user)))


@app.route("/perfect/settings/delete-account", methods=["POST"])
def delete_account():
    username = _api_user()
    if not username:
        return jsonify({"ok": False, "error": "Not logged in"}), 401
    password = (request.json or {}).get("password", "")
    if not password:
        return jsonify({"ok": False, "error": "Password is required"}), 400
    conn = get_db_connection()
    user = conn.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
    if not user:
        conn.close()
        return jsonify({"ok": False, "error": "Account not found"}), 404
    stored_pw = user["password_hash"] or ""
    if stored_pw.startswith(("$2b$", "$2a$", "scrypt:", "pbkdf2:")):
        pw_ok = check_password_hash(stored_pw, password)
    else:
        pw_ok = (stored_pw == password)
    if not pw_ok:
        conn.close()
        return jsonify({"ok": False, "error": "Incorrect password"}), 403
    conn.execute("DELETE FROM users WHERE username=?", (username,))
    conn.commit()
    conn.close()
    session.clear()
    return jsonify({"ok": True})


@app.route("/perfect/api/verify-password", methods=["POST"])
def perfect_verify_password():
    """Check account password — used before sensitive actions like changing diary PIN."""
    username = _api_user()
    if not username:
        return jsonify({"ok": False}), 401
    d = request.get_json(silent=True) or {}
    password = (d.get("password") or "").strip()
    if not password:
        return jsonify({"valid": False}), 400
    conn = get_db_connection()
    row = conn.execute("SELECT password_hash FROM users WHERE username=?", (username,)).fetchone()
    conn.close()
    if not row:
        return jsonify({"valid": False})
    valid = check_password_hash(row["password_hash"], password)
    return jsonify({"valid": valid})


@app.route("/perfect/api/assistant", methods=["POST"])
def perfect_assistant():
    username = _api_user()
    if not username:
        return jsonify({"ok": False}), 401
    d = request.get_json(silent=True) or {}
    message = (d.get("message") or "").strip()
    history  = d.get("history") or []
    if not message:
        return jsonify({"ok": False}), 400

    api_key = os.environ.get("GROQ_API_KEY")
    if not api_key:
        return jsonify({"ok": True, "reply": "AI is not configured yet.", "logged": None})

    is_first_message = len([m for m in history if m.get("role") == "user"]) == 0

    system_prompt = (
        "You are a friendly personal health assistant inside the Too Good app. "
        "Your ONLY output format is a single JSON object -- no extra text before or after.\n\n"

        "LOGGING FIELDS you can capture:\n"
        "  - foods -- what they ate (name, serving size/quantity, calories, protein, carbs, fat)\n"
        "  - weight -- body weight in kg\n"
        "  - steps -- steps walked today\n"
        "  - workout -- exercise description\n"
        "  - hunger -- hunger level today on a scale of 1--10 (1 = not hungry at all, 10 = very hungry)\n"
        "  - energy -- energy level today on a scale of 1--10 (1 = exhausted, 10 = full of energy)\n\n"

        "OUTPUT FORMAT when you have enough info to log:\n"
        '{"reply":"<your message>","logged":{"foods":[{"name":"<name>","serving":"<amount>","calories":<number>,"protein":<number>,"carbs":<number>,"fat":<number>}],"weight":<kg or null>,"steps":<integer or null>,"workout":"<text or null>","hunger":<1-10 or null>,"energy":<1-10 or null>}}\n\n'

        "OUTPUT FORMAT when asking a follow-up or just chatting:\n"
        '{"reply":"<your message>","logged":null}\n\n'

        "RULES -- follow these strictly:\n"
        "1. NEVER guess or estimate portion sizes. If the user mentions food without a quantity "
        "(no grams, no cups, no count, no 'small/medium/large'), ask how much before logging.\n"
        "2. You MAY estimate calories/macros once you have the portion size.\n"
        "3. FIRST USER MESSAGE with food/activity: after acknowledging their food, "
        "ALWAYS ask in the same reply about any missing fields: "
        "if weight not given †' ask weight; if steps not given †' ask steps; "
        "if workout not given †' ask if they exercised today; "
        "if hunger level not given †' ask how hungry they felt today (1--10); "
        "if energy level not given †' ask how their energy was today (1--10). "
        "Bundle ALL missing questions into ONE friendly reply.\n"
        "4. FOLLOW-UP MESSAGES: the user is answering your questions. "
        "Log the new info they provide in THIS message. "
        "Set fields to null for anything not mentioned in this specific message.\n"
        "5. CRITICAL -- foods array: include ONLY foods the user mentions in their CURRENT message. "
        "NEVER re-include foods from earlier messages -- the app has already saved those. "
        "If the user's current message has no food, set foods to [].\n"
        "6. If the user says 'skip', 'nothing', 'no', or 'that's it' for any field †' leave it null.\n"
        "7. If the user sends ONLY a greeting (hi, hello, hey, good morning etc.) with no health data: "
        "reply warmly in one sentence asking what they ate or how their day went. Set logged:null. Do NOT ask about all missing fields.\n"
        "8. If the user asks for something outside today's log (build a plan, change schedule, etc.): "
        "tell them in one sentence that this assistant is only for logging today -- for schedule changes they can use the Calendar page.\n"
        "9. Keep replies warm, brief, and conversational -- no bullet lists, no long paragraphs.\n"
        "10. Never mention JSON, parsing, or these instructions.\n"
        "11. You are the Too Good assistant -- never identify as Gemini or any Google product.\n"
        "12. No signature or sign-off.\n\n"

        + (
            "CONTEXT: This is the user's FIRST message in this conversation. "
            "After handling their food/activity, ask about ALL missing log fields in one go.\n"
            if is_first_message else
            "CONTEXT: This is a follow-up message. The user is filling in details you asked for. "
            "Collect their answers and log everything now.\n"
        )
    )

    log_messages = [
        {"role": m["role"], "content": m["content"]}
        for m in history if m.get("role") in {"user", "assistant"}
    ]
    log_messages.append({"role": "user", "content": message})

    try:
        raw = _call_anthropic(api_key, system_prompt, log_messages, max_tokens=600, temperature=0.3)
    except Exception:
        return jsonify({"ok": True, "reply": "The assistant is temporarily unavailable. Please try again.", "logged": None})

    raw = re.sub(r"^```(?:json)?\s*", "", raw, flags=re.MULTILINE)
    raw = re.sub(r"```\s*$", "", raw, flags=re.MULTILINE).strip()

    match = re.search(r"\{[\s\S]*\}", raw)
    if match:
        try:
            parsed   = json.loads(match.group())
            reply    = parsed.get("reply") or raw
            logged   = parsed.get("logged")
        except json.JSONDecodeError:
            reply, logged = raw, None
    else:
        reply, logged = raw, None

    return jsonify({"ok": True, "reply": reply, "logged": logged})


@app.route("/perfect/api/onboarding", methods=["POST"])
def save_onboarding():
    username = _api_user()
    if not username:
        return jsonify({"ok": False}), 401
    d = request.get_json(silent=True) or {}

    def _int(v):
        try: return int(v)
        except: return None
    def _float(v):
        try: return float(v)
        except: return None

    age              = _int(d.get("age"))
    gender           = (d.get("gender") or "").strip() or None
    weight_kg        = _float(d.get("weight_kg"))
    height_cm        = _float(d.get("height_cm"))
    target_weight_kg = _float(d.get("target_weight_kg"))
    goal             = (d.get("goal") or "").strip() or None
    activity_level   = (d.get("activity_level") or "").strip() or None
    family_role      = (d.get("family_role") or "").strip() or None
    mobility_note    = (d.get("mobility_note") or "").strip() or None
    food_prefs       = (d.get("food_prefs") or "").strip() or None

    conn = get_db_connection()
    conn.execute(
        """UPDATE users SET age=?, gender=?, weight_kg=?, height_cm=?, target_weight_kg=?,
           goal=?, activity_level=?, family_role=?, mobility_note=?, food_prefs=?,
           onboarding_done=1 WHERE username=?""",
        (age,
         b64_encode(gender) if gender else None,
         weight_kg, height_cm, target_weight_kg,
         goal, activity_level, family_role, mobility_note,
         b64_encode(food_prefs) if food_prefs else None,
         username)
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/perfect/api/calendar-edit", methods=["POST"])
def calendar_edit():
    username = _api_user()
    if not username:
        return jsonify({"error": "not logged in"}), 401
    data = request.get_json(silent=True) or {}
    instruction = (data.get("instruction") or "").strip()
    if not instruction:
        return jsonify({"error": "no instruction provided"}), 400

    api_key = os.environ.get("GROQ_API_KEY")
    if not api_key:
        return jsonify({"error": "AI not configured"}), 503

    conn = get_db_connection()
    row = conn.execute(
        "SELECT exercise_types, exercise_days_per_week, rest_day, "
        "session_duration, workout_time_pref, fitness_level FROM users WHERE username=?",
        (username,)
    ).fetchone()

    cur = {
        "exercise_types":         row["exercise_types"] or "not set",
        "exercise_days_per_week": row["exercise_days_per_week"] or "not set",
        "rest_day":               row["rest_day"] or "not set",
        "session_duration":       row["session_duration"] or "not set",
        "workout_time_pref":      row["workout_time_pref"] or "not set",
        "fitness_level":          row["fitness_level"] or "not set",
    }

    prompt = f"""You are a fitness schedule editor. Apply ONLY the changes the user asks for.

CURRENT SCHEDULE:
exercise_types: {cur['exercise_types']}
exercise_days_per_week: {cur['exercise_days_per_week']}
rest_day: {cur['rest_day']}
session_duration: {cur['session_duration']}
workout_time_pref: {cur['workout_time_pref']}
fitness_level: {cur['fitness_level']}

USER INSTRUCTION: "{instruction}"

INTERPRETATION RULES (read carefully before generating output):

TIME RANGES -- if the user gives a time range for any exercise (e.g. "3 to 5", "7pm to 9pm", "9am to 10am"):
  - Calculate start_hour and end_hour EXACTLY as stated. Examples:
      "3 to 5" (no AM/PM, sounds like afternoon) †' start=15, end=17
      "7 to 9 pm" †' start=19, end=21
      "7 to 8" (morning context) †' start=7, end=8
      "9am to 10am" †' start=9, end=10
  - Hour convention (when AM/PM not given): 1--6 = AM, 7--11 = morning, 12--16 = afternoon (add 12), 17--23 = evening.
  - Duration from range: end ' start. Map to nearest: 30min(0.5h), 45min(0.75h), 1hr(1h), 90min(1.5h), 2hr(2h). NEVER default to 2hr.
  - If the user gives DIFFERENT times for DIFFERENT exercises, each gets its own start/end.
  - Put all per-exercise times in "exercise_schedule" (see JSON format below).
  - workout_time_pref: determined from the FIRST exercise's start hour: 0--6†'early_morning, 7--11†'morning, 12--16†'afternoon, 17--23†'evening.

FREQUENCY -- if the user says "everyday", "every day", "daily", "all days", "7 days":
  - exercise_days_per_week = 7
  - rest_day = "" (empty string -- no rest days)

ADDING vs REPLACING exercises:
  - "I want to do [x] everyday/daily", "add [x]", "include [x]", "I also want to do [x]", "I want to do [x] from ... to ..." †' ADD [x] to the existing exercise_types list. KEEP all current exercises. Multiple exercises coexist -- one person can do jogging AND shooting.
  - "only do [x]", "replace everything with [x]", "remove everything and do [x]", "switch entirely to [x]" †' replace the whole list.
  - "remove [x]", "stop doing [x]", "drop [x]" †' remove only [x] from the list, keep everything else.
  - "delete everything", "remove all", "clear everything", "no exercises", "clear my schedule", "delete all" †' set exercise_types to "" (empty string -- clears the entire list).
  - Exercise names: use exactly what the user says. Capitalise properly. Any sport/activity is valid.

REST DAYS:
  - Must be full English day names: Monday, Tuesday, Wednesday, Thursday, Friday, Saturday, Sunday -- comma-separated.
  - If no rest days (exercise every day): use empty string "".
  - exercise_days_per_week = 7 ' number_of_rest_days. (0 rest days ' 7 exercise days.)

OMIT unchanged fields -- ONLY include keys whose values you are changing. The system will keep all omitted fields as they are.

ALLOWED ENUM VALUES (only for fields you are changing):
  session_duration: 30min | 45min | 1hr | 90min | 2hr
  workout_time_pref: early_morning | morning | afternoon | evening | flexible
  fitness_level: beginner | intermediate | advanced | athlete

Return ONLY valid JSON -- no markdown fences, no extra keys, no explanation outside the JSON:
{{
  "exercise_types": "include only if changed",
  "exercise_days_per_week": 0,
  "rest_day": "include only if changed",
  "session_duration": "include only if changed",
  "workout_time_pref": "include only if changed",
  "fitness_level": "include only if changed",
  "exercise_schedule": {{"ExerciseName": {{"start": 15, "end": 17}}}},
  "explanation": "one sentence: exactly what was changed"
}}

exercise_schedule: ONLY include when the user stated specific clock times. Map each exercise name (exactly as it appears in exercise_types) to its start and end hours (integers, 24-hour). If multiple exercises have different times, each gets its own entry. If no times were given, omit this key entirely."""

    try:
        raw = _call_anthropic(api_key, None, [{"role": "user", "content": prompt}], max_tokens=1024, temperature=0.1)
    except Exception:
        conn.close()
        return jsonify({"error": "AI service unavailable. Please try again."}), 502

    # Strip markdown fences if present
    raw = re.sub(r"^```(?:json)?\s*", "", raw, flags=re.MULTILINE)
    raw = re.sub(r"```\s*$", "", raw, flags=re.MULTILINE).strip()
    m = re.search(r"\{[\s\S]*\}", raw)
    if not m:
        conn.close()
        return jsonify({"error": "AI returned unexpected format"}), 502
    try:
        updated = json.loads(m.group())
    except json.JSONDecodeError:
        conn.close()
        return jsonify({"error": "AI returned invalid JSON"}), 502

    def _clean(v, allowed=None):
        s = (v if isinstance(v, str) else "").strip()
        if s.lower() in ("none", "null", ""):
            return None
        return s if (allowed is None or s in allowed) else None

    # Helper: fall back to current DB value if AI omitted/nullified the field
    def _field(ai_val, current_db, allowed=None):
        cleaned = _clean(ai_val, allowed)
        return cleaned if cleaned is not None else (current_db or None)

    # exercise_types: key absent = keep current; key present (even empty) = AI intent (allows clearing all)
    if "exercise_types" not in updated:
        exercise_types = row["exercise_types"]
    else:
        raw_et = updated.get("exercise_types")
        raw_et = (raw_et if isinstance(raw_et, str) else "").strip()
        exercise_types = "" if raw_et.lower() in ("none", "null", "") else raw_et
    session_duration  = _field(updated.get("session_duration"),  row["session_duration"],
                               {"30min","45min","1hr","90min","2hr"})
    workout_time_pref = _field(updated.get("workout_time_pref"), row["workout_time_pref"],
                               {"early_morning","morning","afternoon","evening","flexible"})
    fitness_level     = _field(updated.get("fitness_level"),     row["fitness_level"],
                               {"beginner","intermediate","advanced","athlete"})
    explanation       = (updated.get("explanation") or "Done.").strip()

    # rest_day: key absent = keep current; key present (even empty) = use AI intent
    if "rest_day" not in updated:
        rest_day = row["rest_day"]
    else:
        val = (updated.get("rest_day") or "").strip()
        rest_day = None if val.lower() in ("none", "null", "") else val

    # exercise_days_per_week: allow 1--7 (7 = every day, no rest)
    try:
        ex_days = int(updated.get("exercise_days_per_week", 0))
        if not (1 <= ex_days <= 7): ex_days = None
    except (ValueError, TypeError):
        ex_days = None
    if ex_days is None:
        ex_days = row["exercise_days_per_week"]  # keep current if AI omitted it

    # If every day, clear rest_day
    if ex_days == 7:
        rest_day = None

    # Safety: clamp ex_days so rest_days + ex_days == 7
    if rest_day and ex_days is not None and ex_days != 7:
        n_rest = len([d.strip() for d in rest_day.split(",") if d.strip()])
        if n_rest + ex_days != 7:
            ex_days = max(1, min(7, 7 - n_rest))

    # Build exercise_schedule_json from per-exercise times returned by AI
    import json as _json

    try:
        existing_sched = _json.loads(row["exercise_schedule_json"] or "{}")
    except Exception:
        existing_sched = {}

    # AI returns exercise_schedule: {"ExerciseName": {"start": 15, "end": 17}, ...}
    ai_sched = updated.get("exercise_schedule")
    if isinstance(ai_sched, dict):
        for ex_name, times in ai_sched.items():
            if isinstance(times, dict):
                try:
                    sh = int(times.get("start", times.get("start_hour", 0)))
                    eh = int(times.get("end",   times.get("end_hour",   sh + 1)))
                    if 0 <= sh <= 23 and 0 <= eh <= 24:
                        existing_sched[ex_name.lower().strip()] = {"startH": sh, "endH": eh}
                except (TypeError, ValueError):
                    pass

    exercise_schedule_json = _json.dumps(existing_sched) if existing_sched else None

    # Derive start_hour for response (first exercise's start, for frontend compatibility)
    start_hour = None
    if existing_sched:
        first = next(iter(existing_sched.values()), None)
        if first:
            try: start_hour = int(first["startH"])
            except Exception: pass

    conn.execute(
        "UPDATE users SET exercise_types=?, exercise_days_per_week=?, rest_day=?, "
        "session_duration=?, workout_time_pref=?, fitness_level=?, exercise_schedule_json=? WHERE username=?",
        (exercise_types, ex_days, rest_day, session_duration, workout_time_pref, fitness_level,
         exercise_schedule_json, username)
    )
    conn.commit()
    conn.close()

    return jsonify({
        "ok": True,
        "explanation": explanation,
        "schedule": {
            "exercise_types":         exercise_types or "",
            "exercise_days_per_week": ex_days or 0,
            "rest_day":               rest_day or "",
            "session_duration":       session_duration or "",
            "workout_time_pref":      workout_time_pref or "",
            "fitness_level":          fitness_level or "",
            "workout_start_hour":     start_hour,
            "exercise_schedule":      existing_sched,
        }
    })


@app.route("/perfect/api/save-exercise-times", methods=["POST"])
def save_exercise_times():
    """Save per-exercise schedule times to DB so the nutrition AI can reference them."""
    username = _api_user()
    if not username:
        return jsonify({"ok": False, "error": "not logged in"}), 401
    data = request.get_json(silent=True) or {}
    times_dict = data.get("times", {})  # {exerciseName: {startH, endH}}
    if not isinstance(times_dict, dict):
        return jsonify({"ok": False, "error": "invalid data"}), 400

    import json as _json
    conn = get_db_connection()
    row = conn.execute("SELECT exercise_schedule_json FROM users WHERE username=?", (username,)).fetchone()
    try:
        existing = _json.loads((row["exercise_schedule_json"] if row else None) or "{}")
    except Exception:
        existing = {}

    for ex, times in times_dict.items():
        if isinstance(times, dict) and "startH" in times and "endH" in times:
            existing[ex.lower().strip()] = {"startH": times["startH"], "endH": times["endH"]}

    conn.execute("UPDATE users SET exercise_schedule_json=? WHERE username=?",
                 (_json.dumps(existing), username))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


def _user_context(username):
    """Return common user context dict for page routes."""
    conn = get_db_connection()
    user = _decode_user_row(dict(conn.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()))
    conn.close()
    return dict(
        username=user["username"],
        full_name=user.get("full_name") or "",
        email=user.get("email") or "",
        created_at=(user.get("created_at") or "")[:10],
        age                    = user.get("age") or "",
        gender                 = user.get("gender") or "",
        exercise_types         = user.get("exercise_types") or "",
        exercise_days_per_week = int(user.get("exercise_days_per_week") or 0),
        rest_day               = user.get("rest_day") or "",
        session_duration       = user.get("session_duration") or "",
        workout_time_pref      = user.get("workout_time_pref") or "",
        fitness_level          = user.get("fitness_level") or "",
    )


@app.route("/perfect/exercise")
def perfect_exercise():
    if not session.get("username"):
        return redirect(url_for("login"))
    conn = get_db_connection()
    row = conn.execute(
        "SELECT day_schedule_json FROM users WHERE username=?",
        (session["username"],)
    ).fetchone()
    conn.close()
    day_schedule_json = (row["day_schedule_json"] or "null") if row else "null"
    return render_template(
        "perfect/exercise.html",
        page_active="exercise",
        day_schedule_json=day_schedule_json,
        **_user_context(session["username"])
    )


@app.route("/perfect/api/save-day-schedule", methods=["POST"])
def save_day_schedule():
    username = _api_user()
    if not username:
        return jsonify({"ok": False}), 401
    import json as _json
    data = request.get_json(silent=True) or {}
    sched = data.get("schedule")  # dict: {dayName: {active, exercises: [{name,startH,endH}]}}
    pool  = data.get("pool", [])  # list of exercise names
    if not isinstance(sched, dict):
        return jsonify({"ok": False, "error": "invalid schedule"}), 400
    conn = get_db_connection()
    # Derive exercise_types from pool
    ex_types = ",".join(p.strip() for p in pool if p.strip()) or None
    conn.execute(
        "UPDATE users SET day_schedule_json=?, exercise_types=? WHERE username=?",
        (_json.dumps(sched), ex_types, username)
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/perfect/coach")
def perfect_coach():
    if not session.get("username"):
        return redirect(url_for("login"))
    return render_template("perfect/coach.html", page_active="coach", **_user_context(session["username"]))


@app.route("/perfect/api/coach-chat", methods=["POST"])
def coach_chat():
    username = _api_user()
    if not username:
        return jsonify({"ok": False, "error": "not logged in"}), 401
    data = request.get_json(silent=True) or {}
    user_msg = (data.get("message") or "").strip()[:2000]
    history  = data.get("history", [])
    if not user_msg:
        return jsonify({"ok": False, "error": "empty message"}), 400

    ctx        = _user_context(username)
    user_name  = ctx.get("full_name") or ctx.get("username") or "Champion"
    fit_level  = ctx.get("fitness_level") or "intermediate"
    exercises  = ctx.get("exercise_types") or "general training"

    _conn = get_db_connection()
    _row  = _conn.execute("SELECT goal FROM users WHERE username=?", (username,)).fetchone()
    _conn.close()
    goal = (_row["goal"] if _row and _row["goal"] else None) or "improve fitness"

    sys_prompt = COACH_SYSTEM_PROMPT + f"""

USER PROFILE:
- Name: {user_name}
- Goal: {goal}
- Fitness level: {fit_level}
- Preferred exercises: {exercises}

Address them by name occasionally. Let their goal and level shape your advice naturally."""

    coach_messages = []
    if isinstance(history, list):
        for h in history[-12:]:
            role = "user" if h.get("role") == "user" else "assistant"
            text = str(h.get("text") or "")
            if text:
                coach_messages.append({"role": role, "content": text})
    coach_messages.append({"role": "user", "content": user_msg})

    api_key = os.environ.get("GROQ_API_KEY")
    if not api_key:
        return jsonify({"ok": False, "error": "API not configured"}), 503

    try:
        reply = _call_anthropic(api_key, sys_prompt, coach_messages, max_tokens=450, temperature=0.85)
    except Exception:
        return jsonify({"ok": False, "error": "The coach is temporarily unavailable. Please try again."}), 502

    if not reply:
        return jsonify({"ok": False, "error": "Empty response from AI"}), 502

    return jsonify({"ok": True, "reply": reply})


@app.route("/perfect/calendar")
def perfect_calendar():
    if not session.get("username"):
        return redirect(url_for("login"))
    return render_template("perfect/calendar.html", page_active="calendar", **_user_context(session["username"]))


@app.route("/perfect/adapt")
def perfect_adapt():
    if not session.get("username"):
        return redirect(url_for("login"))
    return render_template("perfect/adapt.html", page_active="adapt", **_user_context(session["username"]))


@app.route("/perfect/log")
def perfect_log():
    if not session.get("username"):
        return redirect(url_for("login"))
    return render_template("perfect/log.html", page_active="log", **_user_context(session["username"]))


@app.route("/perfect/monitor")
def perfect_monitor():
    if not session.get("username"):
        return redirect(url_for("login"))
    return render_template("perfect/monitor.html", page_active="monitor", **_user_context(session["username"]))


MONITOR_SYSTEM_PROMPT = """You are a clinical health analyst AI. Analyze blood test results and give a clear, honest health assessment.

Standard reference ranges (MedlinePlus / ICMR):
CBC: Hemoglobin Male 13.5--17.5 g/dL / Female 12.0--15.5 g/dL | WBC 4,500--11,000 /L | Platelets 150,000--400,000 /L | Hematocrit Male 41--53% / Female 36--46% | RBC Male 4.5--5.9 M/L / Female 4.1--5.1 M/L | MCV 80--100 fL | MCH 27--33 pg
Lipid Panel: Total Cholesterol <200 mg/dL | LDL <100 mg/dL optimal | HDL Male >40 / Female >50 mg/dL | Triglycerides <150 mg/dL
Blood Glucose: Fasting 70--100 mg/dL | HbA1c <5.7% normal / 5.7--6.4% prediabetes / 6.5% diabetes | Postprandial (2h) <140 mg/dL
Liver: ALT 7--56 U/L | AST 10--40 U/L | Total Bilirubin 0.1--1.2 mg/dL | Albumin 3.5--5.0 g/dL | ALP 44--147 U/L | GGT Male 8--61 / Female 5--36 U/L
Kidney: Creatinine Male 0.7--1.3 / Female 0.6--1.1 mg/dL | BUN 7--20 mg/dL | eGFR >60 mL/min/1.73m | Uric Acid Male 3.4--7.0 / Female 2.4--6.0 mg/dL
Thyroid: TSH 0.4--4.0 mIU/L | Free T4 0.8--1.8 ng/dL | Free T3 2.3--4.2 pg/mL
Vitamins & Minerals: Vitamin D 30--100 ng/mL (<20 deficient) | B12 200--900 pg/mL | Iron Male 65--175 / Female 50--170 g/dL | Ferritin Male 24--336 / Female 11--307 ng/mL | Folate 2.7--17.0 ng/mL | Calcium 8.5--10.2 mg/dL | Magnesium 1.7--2.2 mg/dL | Potassium 3.5--5.0 mEq/L | Sodium 136--145 mEq/L
Inflammation: CRP <1.0 mg/L optimal | ESR Male 0--15 / Female 0--20 mm/hr

Respond ONLY in this exact JSON (no markdown fences, no extra text):
{
  "overall_status": "Normal" | "Attention Needed" | "Concerning" | "Critical",
  "assessment": "2-3 honest sentences: what does this report look like overall, any patterns, is anything dangerous",
  "flags": [
    {
      "test": "test name",
      "value": "measured value with unit",
      "status": "low|high|borderline|normal",
      "normal_range": "range string",
      "severity": "mild|moderate|severe",
      "impact": "plain-English one sentence: what this means for health"
    }
  ],
  "warnings": ["any value that needs urgent medical attention -- be direct"],
  "caution": "one-line disclaimer"
}"""

# "€"€ Score / Level System "€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€
MAX_LEVEL = 300

# Tier-based XP per level -- all costs are clean round numbers, total ~20 lakh across 300 levels
_XP_TIERS = [
    (  1,   1,     50),
    (  2,   5,    100),
    (  6,  10,    200),
    ( 11,  25,    400),
    ( 26,  50,    800),
    ( 51,  75,  1_500),
    ( 76, 100,  2_500),
    (101, 150,  4_500),
    (151, 200,  7_000),
    (201, 250, 10_000),
    (251, 300, 16_000),
]
def _xp_for_level(n):
    for s, e, cost in _XP_TIERS:
        if s <= n <= e:
            return cost
    return 16_000

# Build cumulative threshold table once at startup
# LEVEL_TABLE[n] = total XP required to REACH level (n+1)  (index 0 †' reach level 1 = 0)
def _build_level_table():
    table = [0]
    for n in range(1, MAX_LEVEL):          # transitions level 1†'2 ... 299†'300
        table.append(table[-1] + _xp_for_level(n))
    return table

import bisect as _bisect
LEVEL_TABLE = _build_level_table()        # len = 300

def _compute_level(total_xp):
    """Return (level 1-300, xp_in_current_level, xp_needed_for_this_level)."""
    total_xp = max(0, total_xp)
    i        = min(MAX_LEVEL - 1, _bisect.bisect_right(LEVEL_TABLE, total_xp) - 1)
    level    = i + 1
    xp_start = LEVEL_TABLE[i]
    xp_need  = _xp_for_level(level) if level < MAX_LEVEL else _xp_for_level(MAX_LEVEL)
    return level, round(total_xp - xp_start, 2), xp_need

def _get_xp_multiplier(level):
    """XP multiplier based on current level tier."""
    if level >= 200: return 20
    if level >= 100: return 5
    return 1

SCORE_ACTIONS = {
    "login":        {"xp": 5,  "daily_limit": 1,  "label": "Daily check-in"},
    "food_log":     {"xp": 8,  "daily_limit": 1,  "label": "Logged food"},
    "calorie_goal": {"xp": 12, "daily_limit": 1,  "label": "Calorie goal achieved"},
    "exercise":     {"xp": 8,  "daily_limit": 1,  "label": "Exercise logged"},
    "blood_report": {"xp": 100, "daily_limit": 999, "monthly_limit": 1, "label": "Blood report analyzed"},
}

# Streak milestone bonuses (base -- also multiplied by level multiplier)
STREAK_BONUSES = {100: 70, 30: 50, 7: 5}

# "€"€ Achievement definitions "€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€
ACHIEVEMENT_META = {
    "first_checkin":  {"emoji": "", "title": "Dawn Patrol",        "desc": "First daily check-in"},
    "first_workout":  {"emoji": "", "title": "Off the Couch",      "desc": "First workout logged"},
    "first_blood":    {"emoji": "", "title": "Lab Rat",            "desc": "First blood report analyzed"},
    "streak_7":       {"emoji": """, "title": "On Fire",            "desc": "7-day streak"},
    "streak_30":      {"emoji": "'", "title": "Diamond Discipline", "desc": "30-day streak"},
    "streak_100":     {"emoji": "''", "title": "Centurion",          "desc": "100-day streak"},
    "level_10":       {"emoji": "", "title": "Getting Serious",    "desc": "Reached level 10"},
    "level_50":       {"emoji": "", "title": "Power User",         "desc": "Reached level 50"},
    "level_100":      {"emoji": "", "title": "Elite",              "desc": "Reached level 100"},
    "xp_1000":        {"emoji": "'", "title": "XP Earner",          "desc": "Earned 1,000 XP total"},
    "xp_10000":       {"emoji": "'", "title": "XP Machine",         "desc": "Earned 10,000 XP total"},
    "xp_100000":      {"emoji": "", "title": "XP Legend",          "desc": "Earned 100,000 XP total"},
    "workouts_10":    {"emoji": "'", "title": "Getting Strong",     "desc": "10 workouts logged"},
    "workouts_50":    {"emoji": "", "title": "Half Century",       "desc": "50 workouts logged"},
    "workouts_100":   {"emoji": """, "title": "Iron Will",          "desc": "100 workouts logged"},
    "buddy_1":        {"emoji": "", "title": "Social Starter",     "desc": "First buddy added"},
    "buddy_5":        {"emoji": "", "title": "Social Butterfly",   "desc": "5 buddies"},
    "club_join":      {"emoji": "", "title": "Team Player",        "desc": "Joined a club"},
    "challenge_done": {"emoji": "-", "title": "Challenge Accepted", "desc": "Completed a challenge"},
}

def _check_achievements(conn, username):
    """Check milestones and award any newly unlocked achievements. Returns list of new badge keys."""
    already = {r["badge"] for r in conn.execute(
        "SELECT badge FROM achievements WHERE username=?", (username,)
    ).fetchall()}
    new_badges = []

    def award(badge):
        if badge not in already:
            conn.execute("INSERT OR IGNORE INTO achievements (username, badge) VALUES (?,?)", (username, badge))
            already.add(badge)
            new_badges.append(badge)

    user = conn.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
    if not user:
        return []

    streak   = user["score_streak"] or 0
    total_xp = user["score_xp"]     or 0
    level, _, _ = _compute_level(total_xp)

    if streak >= 7:   award("streak_7")
    if streak >= 30:  award("streak_30")
    if streak >= 100: award("streak_100")
    if level  >= 10:  award("level_10")
    if level  >= 50:  award("level_50")
    if level  >= 100: award("level_100")
    if total_xp >= 1_000:    award("xp_1000")
    if total_xp >= 10_000:   award("xp_10000")
    if total_xp >= 100_000:  award("xp_100000")

    workouts = conn.execute(
        "SELECT COUNT(*) FROM score_events WHERE username=? AND event_type='exercise'", (username,)
    ).fetchone()[0]
    if workouts >= 1:   award("first_workout")
    if workouts >= 10:  award("workouts_10")
    if workouts >= 50:  award("workouts_50")
    if workouts >= 100: award("workouts_100")

    if conn.execute("SELECT COUNT(*) FROM score_events WHERE username=? AND event_type='login'", (username,)).fetchone()[0] >= 1:
        award("first_checkin")
    if conn.execute("SELECT COUNT(*) FROM score_events WHERE username=? AND event_type='blood_report'", (username,)).fetchone()[0] >= 1:
        award("first_blood")

    buddies = conn.execute(
        "SELECT COUNT(*) FROM buddies WHERE (requester=? OR recipient=?) AND status='accepted'", (username, username)
    ).fetchone()[0]
    if buddies >= 1: award("buddy_1")
    if buddies >= 5: award("buddy_5")

    if conn.execute("SELECT COUNT(*) FROM club_members WHERE username=?", (username,)).fetchone()[0] >= 1:
        award("club_join")

    conn.commit()
    return new_badges


def _seed_monthly_challenges():
    """Auto-create challenges for the current month if they don't exist yet."""
    pk         = datetime.utcnow().strftime("%Y-%m")
    month_name = datetime.utcnow().strftime("%B %Y")
    defaults   = [
        ("Workout Warrior",  f"Log 20 workouts in {month_name}",           "workouts",  20),
        ("XP Grind",         f"Earn 5,000 XP in {month_name}",             "xp",      5000),
        ("Daily Devotion",   f"Check in every day this month -- {month_name}", "checkins", 20),
        ("Food Journal",     f"Log food 25 times in {month_name}",          "food_logs", 25),
    ]
    conn = get_db_connection()
    for title, desc, metric, target in defaults:
        if not conn.execute("SELECT 1 FROM challenges WHERE metric=? AND period_key=?", (metric, pk)).fetchone():
            conn.execute(
                "INSERT INTO challenges (title,description,metric,target,period,period_key) VALUES (?,?,?,?,?,?)",
                (title, desc, metric, target, "monthly", pk)
            )
    conn.commit()
    conn.close()

_seed_monthly_challenges()


def _challenge_progress(conn, username, challenge):
    """Compute how much progress a user has on a challenge this period."""
    pk     = challenge["period_key"]
    metric = challenge["metric"]
    if metric == "workouts":
        val = conn.execute(
            "SELECT COUNT(*) FROM score_events WHERE username=? AND event_type='exercise' AND strftime('%Y-%m',awarded_at)=?",
            (username, pk)
        ).fetchone()[0]
    elif metric == "xp":
        val = conn.execute(
            "SELECT COALESCE(SUM(xp_awarded),0) FROM score_events WHERE username=? AND strftime('%Y-%m',awarded_at)=?",
            (username, pk)
        ).fetchone()[0]
    elif metric == "checkins":
        val = conn.execute(
            "SELECT COUNT(*) FROM score_events WHERE username=? AND event_type='login' AND strftime('%Y-%m',awarded_at)=?",
            (username, pk)
        ).fetchone()[0]
    elif metric == "food_logs":
        val = conn.execute(
            "SELECT COUNT(*) FROM score_events WHERE username=? AND event_type='food_log' AND strftime('%Y-%m',awarded_at)=?",
            (username, pk)
        ).fetchone()[0]
    else:
        val = 0
    return int(val)


def _award_xp(username, action):
    """Award XP for an action. Returns dict with awarded, xp_gained, new totals."""
    cfg = SCORE_ACTIONS.get(action)
    if not cfg:
        return {"awarded": False, "reason": "unknown action"}

    today_str = datetime.utcnow().strftime("%Y-%m-%d")
    conn      = get_db_connection()

    # Check monthly limit (for actions like blood_report)
    if cfg.get("monthly_limit"):
        month_str = datetime.utcnow().strftime("%Y-%m")
        count_month = conn.execute(
            "SELECT COUNT(*) FROM score_events WHERE username=? AND event_type=? AND strftime('%Y-%m', awarded_at)=?",
            (username, action, month_str)
        ).fetchone()[0]
        if count_month >= cfg["monthly_limit"]:
            conn.close()
            return {"awarded": False, "reason": "monthly limit reached"}

    # Check daily limit
    count_today = conn.execute(
        "SELECT COUNT(*) FROM score_events WHERE username=? AND event_type=? AND DATE(awarded_at)=?",
        (username, action, today_str)
    ).fetchone()[0]
    if count_today >= cfg["daily_limit"]:
        conn.close()
        return {"awarded": False, "reason": "daily limit reached"}

    # Get current level for multiplier
    cur_xp   = conn.execute("SELECT COALESCE(score_xp,0) FROM users WHERE username=?", (username,)).fetchone()[0]
    cur_lvl, _, _ = _compute_level(cur_xp)
    mult     = _get_xp_multiplier(cur_lvl)

    xp_base   = cfg["xp"]
    streak_bonus = 0
    streak    = 0

    if action == "login":
        row       = conn.execute("SELECT score_streak, score_last_date FROM users WHERE username=?", (username,)).fetchone()
        streak    = row["score_streak"] or 0
        last_date = row["score_last_date"] or ""
        yesterday = (datetime.utcnow().date() - timedelta(days=1)).isoformat()
        if last_date == yesterday:
            streak += 1
        elif last_date != today_str:
            streak = 1
        for threshold in sorted(STREAK_BONUSES, reverse=True):
            if streak >= threshold:
                streak_bonus = STREAK_BONUSES[threshold]
                break
        conn.execute("UPDATE users SET score_streak=?, score_last_date=? WHERE username=?",
                     (streak, today_str, username))

    total_xp_gained = round((xp_base + streak_bonus) * mult, 2)

    notes = []
    if mult > 1:       notes.append(f"{mult}-- level bonus")
    if streak_bonus:   notes.append(f"+{streak_bonus} streak")
    note = ", ".join(notes) or None

    conn.execute(
        "INSERT INTO score_events (username, event_type, xp_awarded, note) VALUES (?,?,?,?)",
        (username, action, total_xp_gained, note)
    )
    conn.execute("UPDATE users SET score_xp = COALESCE(score_xp,0) + ? WHERE username=?",
                 (total_xp_gained, username))
    conn.commit()

    new_xp = conn.execute("SELECT score_xp FROM users WHERE username=?", (username,)).fetchone()["score_xp"]
    new_badges = _check_achievements(conn, username)
    conn.close()

    level, xp_in, xp_needed = _compute_level(new_xp)
    return {
        "awarded":           True,
        "xp_gained":         total_xp_gained,
        "streak_bonus":      streak_bonus,
        "multiplier":        mult,
        "total_xp":          new_xp,
        "level":             level,
        "new_achievements":  [{"badge": b, **ACHIEVEMENT_META.get(b, {"emoji":"","title":b,"desc":""})} for b in new_badges],
        "xp_in_level":  xp_in,
        "xp_per_level": xp_needed,
    }


@app.route("/perfect/score")
def perfect_score():
    if not session.get("username"):
        return redirect(url_for("login"))
    return render_template("perfect/score.html", page_active="score", **_user_context(session["username"]))


@app.route("/perfect/api/daily-logs", methods=["GET", "POST"])
def perfect_daily_logs():
    username = _api_user()
    if not username:
        return jsonify({"error": "Not logged in"}), 401
    conn = get_db_connection()
    if request.method == "GET":
        rows = conn.execute(
            "SELECT date, data FROM daily_logs WHERE username=? ORDER BY date DESC LIMIT 365",
            (username,)
        ).fetchall()
        conn.close()
        return jsonify({"ok": True, "logs": [{"date": r["date"], **json.loads(r["data"])} for r in rows]})
    # POST — upsert a batch of logs
    body = request.get_json(silent=True) or {}
    logs = body.get("logs", [])
    for entry in logs:
        date = entry.get("date")
        if not date:
            continue
        data = json.dumps({k: v for k, v in entry.items() if k != "date"})
        conn.execute(
            "INSERT INTO daily_logs (username, date, data, updated_at) VALUES (?,?,?,CURRENT_TIMESTAMP) "
            "ON CONFLICT(username, date) DO UPDATE SET data=excluded.data, updated_at=CURRENT_TIMESTAMP",
            (username, date, data)
        )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/perfect/api/score", methods=["GET"])
def perfect_score_api_get():
    username = _api_user()
    if not username:
        return jsonify({"error": "Not logged in"}), 401
    conn     = get_db_connection()
    user     = conn.execute("SELECT score_xp, score_streak, score_last_date FROM users WHERE username=?",
                            (username,)).fetchone()
    events   = conn.execute(
        "SELECT event_type, xp_awarded, awarded_at, note FROM score_events WHERE username=? ORDER BY awarded_at DESC LIMIT 30",
        (username,)
    ).fetchall()
    total_xp = user["score_xp"] or 0
    my_rank  = conn.execute(
        "SELECT COUNT(*)+1 FROM users WHERE COALESCE(score_xp,0) > ?", (total_xp,)
    ).fetchone()[0]
    conn.close()

    streak   = user["score_streak"] or 0
    level, xp_in, xp_needed = _compute_level(total_xp)

    return jsonify({
        "ok": True,
        "total_xp":    round(total_xp, 2),
        "level":       level,
        "max_level":   MAX_LEVEL,
        "xp_in_level": round(xp_in, 2),
        "xp_per_level": xp_needed,
        "streak":      streak,
        "multiplier":  _get_xp_multiplier(level),
        "my_rank":     my_rank,
        "events":      [dict(e) for e in events],
    })


@app.route("/perfect/api/score/award", methods=["POST"])
def perfect_score_award():
    username = _api_user()
    if not username:
        return jsonify({"error": "Not logged in"}), 401
    data   = request.get_json(force=True) or {}
    action = data.get("action", "")
    result = _award_xp(username, action)
    return jsonify(result)


@app.route("/perfect/api/leaderboard", methods=["GET"])
def perfect_leaderboard():
    me = _api_user()
    if not me:
        return jsonify({"error": "Not logged in"}), 401
    conn = get_db_connection()

    # Top 10 by XP
    rows = conn.execute(
        "SELECT username, full_name, COALESCE(score_xp,0) AS xp, COALESCE(score_streak,0) AS streak "
        "FROM users ORDER BY score_xp DESC LIMIT 10"
    ).fetchall()

    # My rank (number of users with strictly more XP than me, +1)
    my_xp = conn.execute(
        "SELECT COALESCE(score_xp,0) FROM users WHERE username=?", (me,)
    ).fetchone()[0]
    my_rank = conn.execute(
        "SELECT COUNT(*)+1 FROM users WHERE COALESCE(score_xp,0) > ?", (my_xp,)
    ).fetchone()[0]
    total_users = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    conn.close()

    board = []
    for i, r in enumerate(rows):
        try:    display = b64_decode(r["full_name"]) if r["full_name"] else r["username"]
        except: display = r["username"]
        display = (display[:18] + "...") if len(display) > 19 else display
        lvl, _, _ = _compute_level(r["xp"])
        board.append({
            "rank":     i + 1,
            "name":     display,
            "is_me":    r["username"] == me,
            "level":    lvl,
            "xp":       round(r["xp"]),
            "streak":   r["streak"],
        })

    # If current user is not in top 10, append their own entry
    me_in_top = any(e["is_me"] for e in board)
    if not me_in_top:
        my_lvl, _, _ = _compute_level(my_xp)
        board.append({
            "rank":   my_rank,
            "name":   "You",
            "is_me":  True,
            "level":  my_lvl,
            "xp":     round(my_xp),
            "streak": 0,
            "gap":    True,
        })

    return jsonify({"ok": True, "board": board, "my_rank": my_rank, "total_users": total_users})


# "€"€ End Score System "€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€

EXTRACT_SYSTEM_PROMPT = """You are a medical OCR engine. Read the blood test report image and extract EVERY numeric test result visible anywhere in the image.

CRITICAL RULES:
- Read the ENTIRE image. Blood reports have multiple sections: CBC, Lipid Profile, LFT, KFT, TFT, Vitamins, etc. Extract from ALL sections.
- A typical blood report has 15 to 40 values. Extract as many as you can find.
- DO NOT stop after the first value or first section.
- Values must be numeric only. Strip units (g/dL, U/L, mg/dL, etc.), reference ranges, and H/L flags. Just the number.
- Use the EXACT field ID keys listed below. Match by test name similarity.

COMMON NAME MAPPINGS (test name on report †' field ID):
Hb / Haemoglobin / Hemoglobin †' hemoglobin
WBC / Total Leucocyte Count / TLC †' wbc
RBC / Red Cell Count †' rbc
Platelets / Platelet Count / PLT †' platelets
PCV / Hematocrit / HCT †' hematocrit
MCV †' mcv
MCH †' mch
Total Cholesterol / T.Chol †' total_cholesterol
LDL / LDL-C / LDL Cholesterol †' ldl
HDL / HDL-C / HDL Cholesterol †' hdl
Triglycerides / TG / VLDL (--5) ' triglycerides
Fasting Blood Sugar / FBS / FBG / Glucose Fasting †' fasting_glucose
HbA1c / Glycated Haemoglobin †' hba1c
Post Prandial / PP Glucose / Postprandial †' postprandial
SGPT / ALT †' alt
SGOT / AST †' ast
Total Bilirubin / S.Bilirubin †' bilirubin_total
Albumin / S.Albumin †' albumin
ALP / Alkaline Phosphatase †' alp
GGT / Gamma GT †' ggt
Creatinine / S.Creatinine †' creatinine
BUN / Blood Urea Nitrogen / Urea †' bun
eGFR / GFR †' egfr
Uric Acid / S.Uric Acid †' uric_acid
TSH / Thyroid Stimulating Hormone †' tsh
Free T4 / FT4 †' free_t4
Free T3 / FT3 †' free_t3
Vitamin D / 25-OH Vitamin D / 25(OH)D †' vitamin_d
Vitamin B12 / Cobalamin †' vitamin_b12
Serum Iron / S.Iron †' iron
Ferritin / S.Ferritin †' ferritin
Folate / Folic Acid †' folate
Calcium / S.Calcium †' calcium
Magnesium / S.Magnesium †' magnesium
Potassium / S.Potassium †' potassium
Sodium / S.Sodium †' sodium
CRP / C-Reactive Protein †' crp
ESR †' esr

Return ONLY a JSON object with field_id: "numeric_value" pairs. No extra text.
Example of a GOOD complete response: {"hemoglobin":"12.5","wbc":"6.2","rbc":"4.1","platelets":"210","hematocrit":"38","mcv":"88","mch":"28","total_cholesterol":"195","ldl":"120","hdl":"42","triglycerides":"160","alt":"32","ast":"28","tsh":"2.1","vitamin_d":"18.4","creatinine":"0.9"}"""


@app.route("/perfect/api/monitor/extract", methods=["POST"])
def perfect_monitor_extract():
    if not _api_user():
        return jsonify({"error": "Not logged in"}), 401
    api_key = os.environ.get("GROQ_API_KEY")
    if not api_key:
        return jsonify({"error": "AI not configured"}), 503

    data       = request.get_json(force=True) or {}
    image_b64  = data.get("image_b64")
    image_mime = data.get("image_mime", "image/jpeg")
    if not image_b64:
        return jsonify({"error": "No image provided"}), 400
    if len(image_b64) > 14_000_000:
        return jsonify({"error": "Image too large. Please use an image under 10 MB."}), 400

    # Try Groq vision models in order until one works
    vision_models = [
        "meta-llama/llama-4-scout-17b-16e-instruct",
        "meta-llama/llama-4-maverick-17b-128e-instruct",
        "llama-3.2-90b-vision-preview",
    ]
    EXTRACTION_TEXT = (
        "This is a medical laboratory blood test report. You MUST extract EVERY numeric test result visible in this image.\n\n"
        "STEP 1 -- Scan the entire image from top to bottom. Blood reports have multiple sections: CBC, Lipid, LFT, KFT, TFT, Vitamins, Inflammation. Read ALL of them.\n"
        "STEP 2 -- For each row that has a test name and a numeric result, record the numeric value (digits only, no units, no ranges).\n"
        "STEP 3 -- Map each test name to the correct field_id below and return a single JSON object.\n\n"
        "Field IDs to use (EXACT keys):\n"
        "hemoglobin, wbc, rbc, platelets, hematocrit, mcv, mch,\n"
        "total_cholesterol, ldl, hdl, triglycerides,\n"
        "fasting_glucose, hba1c, postprandial,\n"
        "alt, ast, bilirubin_total, albumin, alp, ggt,\n"
        "creatinine, bun, egfr, uric_acid,\n"
        "tsh, free_t4, free_t3,\n"
        "vitamin_d, vitamin_b12, iron, ferritin, folate, calcium, magnesium, potassium, sodium,\n"
        "crp, esr\n\n"
        "Common mappings: Hb/Haemoglobin†'hemoglobin, TLC/WBC†'wbc, PCV†'hematocrit, SGPT†'alt, SGOT†'ast, "
        "S.Bilirubin†'bilirubin_total, S.Creatinine†'creatinine, S.Uric Acid†'uric_acid, "
        "T.Cholesterol†'total_cholesterol, 25(OH)D†'vitamin_d, Cobalamin†'vitamin_b12.\n\n"
        "Return ONLY a JSON object. Example: {\"hemoglobin\":\"12.5\",\"wbc\":\"7.2\",\"platelets\":\"210\",\"alt\":\"32\"}\n"
        "A typical blood report has 15--35 values. Extract as many as you can find."
    )
    user_content = [
        {"type": "image_url", "image_url": {"url": f"data:{image_mime};base64,{image_b64}"}},
        {"type": "text", "text": EXTRACTION_TEXT}
    ]

    raw = None
    last_err = ""
    for vmodel in vision_models:
        vision_payload = {
            "model": vmodel,
            "max_tokens": 2048,
            "temperature": 0.0,
            "messages": [
                {"role": "system", "content": EXTRACT_SYSTEM_PROMPT},
                {"role": "user", "content": user_content}
            ]
        }
        req = Request(
            GROQ_API_URL,
            data=json.dumps(vision_payload).encode("utf-8"),
            headers={
                "Content-Type":  "application/json",
                "Authorization": f"Bearer {api_key}",
                "User-Agent":    "groq-python/0.9.0",
            },
            method="POST",
        )
        try:
            with urlopen(req, timeout=60) as resp:
                result = json.loads(resp.read().decode("utf-8"))
            raw = result["choices"][0]["message"]["content"].strip()
            break
        except HTTPError as err:
            last_err = err.read().decode("utf-8", errors="replace")
        except Exception as err:
            last_err = str(err)

    if raw is None:
        return jsonify({"error": "Image analysis service unavailable. Please try again."}), 502

    # Strip markdown fences
    raw = re.sub(r'^```[a-zA-Z]*\n?', '', raw).rstrip('`').strip()

    # Try direct parse, then regex search for first {...} block
    try:
        values = json.loads(raw)
    except Exception:
        m = re.search(r'\{[^{}]*\}', raw, re.DOTALL)
        if m:
            try:
                values = json.loads(m.group())
            except Exception:
                return jsonify({"error": "Could not parse report values", "raw": raw[:300]}), 500
        else:
            return jsonify({"error": "No values found in image", "raw": raw[:300]}), 500

    known = {
        "hemoglobin","wbc","rbc","platelets","hematocrit","mcv","mch",
        "total_cholesterol","ldl","hdl","triglycerides",
        "fasting_glucose","hba1c","postprandial",
        "alt","ast","bilirubin_total","albumin","alp","ggt",
        "creatinine","bun","egfr","uric_acid",
        "tsh","free_t4","free_t3",
        "vitamin_d","vitamin_b12","iron","ferritin","folate",
        "calcium","magnesium","potassium","sodium","crp","esr"
    }
    clean = {k: str(v) for k, v in values.items() if k in known and v not in (None, "", "null")}
    return jsonify({"ok": True, "values": clean})


@app.route("/perfect/api/monitor/save", methods=["POST"])
def perfect_monitor_save():
    username = _api_user()
    if not username:
        return jsonify({"error": "Not logged in"}), 401
    data = request.get_json(force=True) or {}
    report_json = data.get("report")
    label       = data.get("label", "")
    if not report_json:
        return jsonify({"error": "No report data"}), 400
    report_str = json.dumps(report_json)
    conn = get_db_connection()
    conn.execute(
        "UPDATE users SET blood_report_json = ? WHERE username = ?",
        (report_str, username)
    )
    conn.execute(
        "INSERT INTO blood_scan_history (username, label, result_json) VALUES (?, ?, ?)",
        (username, label, report_str)
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/perfect/api/monitor/history", methods=["GET"])
def perfect_monitor_history():
    username = _api_user()
    if not username:
        return jsonify({"error": "Not logged in"}), 401
    conn = get_db_connection()
    rows = conn.execute(
        "SELECT id, scanned_at, label, result_json FROM blood_scan_history "
        "WHERE username = ? ORDER BY scanned_at DESC LIMIT 50",
        (username,)
    ).fetchall()
    conn.close()
    history = []
    for row in rows:
        try:
            result = json.loads(row["result_json"])
        except Exception:
            result = {}
        history.append({
            "id":         row["id"],
            "scanned_at": row["scanned_at"],
            "label":      row["label"] or "",
            "result":     result,
        })
    return jsonify({"ok": True, "history": history})


@app.route("/perfect/api/monitor/clear", methods=["POST"])
def perfect_monitor_clear():
    username = _api_user()
    if not username:
        return jsonify({"error": "Not logged in"}), 401
    conn = get_db_connection()
    conn.execute(
        "UPDATE users SET blood_report_json = NULL WHERE username = ?",
        (username,)
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/perfect/api/monitor", methods=["POST"])
def perfect_monitor_api():
    _monitor_user = _api_user()
    if not _monitor_user:
        return jsonify({"error": "Not logged in"}), 401

    api_key = os.environ.get("GROQ_API_KEY")
    if not api_key:
        return jsonify({"error": "AI not configured"}), 503

    data       = request.get_json(force=True) or {}
    image_b64  = data.get("image_b64")
    image_mime = data.get("image_mime", "image/jpeg")
    values     = data.get("values", {})
    gender     = data.get("gender", "unspecified")
    age        = data.get("age", "unspecified")

    if image_b64:
        if len(image_b64) > 14_000_000:
            return jsonify({"error": "Image too large. Please use an image under 10 MB."}), 400
        user_content = [
            {"type": "image", "source": {"type": "base64", "media_type": image_mime, "data": image_b64}},
            {"type": "text", "text": f"Patient gender: {gender}, age: {age}. Analyze this blood report."}
        ]
    else:
        values_text = "\n".join(f"- {k}: {v}" for k, v in values.items() if str(v).strip())
        if not values_text:
            return jsonify({"error": "No values provided"}), 400
        user_content = f"Patient gender: {gender}, age: {age}.\n\nBlood test results:\n{values_text}"

    try:
        raw = _call_anthropic(api_key, MONITOR_SYSTEM_PROMPT,
                              [{"role": "user", "content": user_content}],
                              max_tokens=2048, temperature=0.0)
    except Exception:
        return jsonify({"error": "Analysis service unavailable. Please try again."}), 502

    raw = re.sub(r'^```[a-zA-Z]*\n?', '', raw).rstrip('`').strip()

    def _deterministic_status(flags):
        severities = {f.get("severity","") for f in flags if f.get("status") not in ("normal",None,"")}
        abnormal   = {f.get("status","")   for f in flags if f.get("status") not in ("normal",None,"")}
        if "severe"   in severities: return "Critical"
        if "moderate" in severities: return "Concerning"
        if abnormal:                 return "Attention Needed"
        return "Normal"

    try:
        parsed = json.loads(raw)
        parsed["overall_status"] = _deterministic_status(parsed.get("flags", []))
        _award_xp(_monitor_user, "blood_report")
        return jsonify({"ok": True, "result": parsed})
    except Exception:
        m = re.search(r'\{[\s\S]*\}', raw)
        if m:
            try:
                parsed = json.loads(m.group())
                parsed["overall_status"] = _deterministic_status(parsed.get("flags", []))
                _award_xp(_monitor_user, "blood_report")
                return jsonify({"ok": True, "result": parsed})
            except Exception:
                pass
        return jsonify({"ok": True, "result": {"raw": raw}})


@app.route("/perfect")
def perfect_home():
    if not session.get("username"):
        return redirect(url_for("login"))
    conn = get_db_connection()
    user = _decode_user_row(dict(conn.execute("SELECT * FROM users WHERE username = ?", (session["username"],)).fetchone()))
    conn.close()
    family_holiday     = _get_family_holiday(user.get("family_role"))
    country_festivals  = _get_country_festivals(user.get("country") or "")
    return render_template(
        "perfect/index.html",
        page_active="dashboard",
        username=user["username"],
        full_name=user.get("full_name") or "",
        email=user.get("email") or "",
        created_at=(user.get("created_at") or "")[:10],
        ban_until=_resolve_ban(user["username"]),
        family_holiday=family_holiday,
        country_festivals=country_festivals,
        exercise_types         = user.get("exercise_types") or "",
        exercise_days_per_week = int(user.get("exercise_days_per_week") or 0),
        rest_day               = user.get("rest_day") or "",
        session_duration       = user.get("session_duration") or "",
        workout_time_pref      = user.get("workout_time_pref") or "",
        fitness_level          = user.get("fitness_level") or "",
        onboarding_needed      = not user.get("onboarding_done"),
    )


@app.route("/signup", methods=["GET", "POST"])
def signup():
    if request.method == "GET":
        return render_template("signup.html", active_page="signup")

    ip = request.remote_addr or "unknown"
    if _rate_limit_exceeded(f"register:{ip}", max_attempts=5, window_secs=3600):
        return render_template("signup.html", active_page="signup",
            server_error_field="username",
            server_error_msg="Too many registration attempts. Please try again later.")

    username  = request.form.get("username", "").strip()
    password  = request.form.get("password", "")
    confirm   = request.form.get("confirm_password", "")
    full_name = request.form.get("full_name", "").strip()
    country   = request.form.get("country", "").strip()
    email     = request.form.get("email", "").strip().lower()

    prev = dict(prev_username=username, prev_full_name=full_name, prev_country=country, prev_email=email)

    if not email or "@" not in email or "." not in email.split("@")[-1]:
        return render_template("signup.html", active_page="signup",
            server_error_field="email",
            server_error_msg="Please enter a valid email address.", **prev)

    if not USERNAME_RE.match(username):
        return render_template("signup.html", active_page="signup",
            server_error_field="username",
            server_error_msg="Only letters (a--z), numbers, and underscores -- no spaces.", **prev)

    if len(password) < 8:
        return render_template("signup.html", active_page="signup",
            server_error_field="password",
            server_error_msg="Password must be at least 8 characters.", **prev)

    if password != confirm:
        return render_template("signup.html", active_page="signup",
            server_error_field="confirm",
            server_error_msg="Passwords do not match.", **prev)

    conn = get_db_connection()
    try:
        conn.execute(
            "INSERT INTO users (username, password_hash, email, full_name, country) VALUES (?, ?, ?, ?, ?)",
            (username, generate_password_hash(password), b64_encode(email) or None,
             b64_encode(full_name) or None, b64_encode(country) or None),
        )
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return render_template("signup.html", active_page="signup",
            server_error_field="username",
            server_error_msg="That username is already taken. Please choose another.", **prev)

    user = conn.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
    conn.close()

    session.clear()
    session["username"] = username
    session["user_id"]  = user["id"]
    return redirect(url_for("perfect_home") + "?msg=created")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        return render_template("login.html", active_page="login")

    username = request.form.get("username", "").strip()
    password = request.form.get("password", "")
    ip       = request.remote_addr or "unknown"
    rl_key   = f"login:{ip}:{username}"
    if _rate_limit_exceeded(rl_key, max_attempts=10, window_secs=300):
        return render_template(
            "login.html", active_page="login",
            prev_username=username,
            server_error_field="password",
            server_error_msg="Too many login attempts. Please try again in a few minutes.",
        )

    conn = get_db_connection()
    user = conn.execute(
        "SELECT * FROM users WHERE username = ?", (username,)
    ).fetchone()

    if user is None or not user["password_hash"]:
        conn.close()
        return render_template(
            "login.html", active_page="login",
            prev_username=username,
            server_error_field="password",
            server_error_msg="Incorrect username or password.",
        )

    stored = user["password_hash"]
    if stored.startswith(("$2b$", "$2a$", "scrypt:", "pbkdf2:")):
        match = check_password_hash(stored, password)
    else:
        match = (stored == password)

    if not match:
        conn.close()
        return render_template(
            "login.html", active_page="login",
            prev_username=username,
            server_error_field="password",
            server_error_msg="Incorrect username or password.",
        )

    if not stored.startswith(("$2b$", "$2a$", "scrypt:", "pbkdf2:")):
        conn.execute("UPDATE users SET password_hash=? WHERE username=?",
                     (generate_password_hash(password), user["username"]))
        conn.commit()

    conn.close()
    session.clear()
    session["username"] = user["username"]
    session["user_id"]  = user["id"]
    return redirect(url_for("perfect_home") + "?msg=login")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("home"))


# "€"€ Google OAuth "€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€

@app.route("/auth/google")
def google_login():
    if not os.environ.get("GOOGLE_CLIENT_ID"):
        flash("Google sign-in is not configured. Add GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET to your .env file.")
        return redirect(url_for("signup"))
    redirect_uri = url_for("google_callback", _external=True)
    return google.authorize_redirect(redirect_uri)


@app.route("/auth/google/callback")
def google_callback():
    try:
        token = google.authorize_access_token()
    except Exception:
        flash("Google sign-in was cancelled or failed. Please try again.")
        return redirect(url_for("signup"))

    user_info = token.get("userinfo")
    if not user_info:
        flash("Could not retrieve your Google account info. Please try again.")
        return redirect(url_for("signup"))

    google_id = user_info["sub"]
    email = user_info.get("email", "")
    name = user_info.get("name", "")

    conn = get_db_connection()

    # Already has a Google-linked account -- log in
    user = conn.execute(
        "SELECT * FROM users WHERE google_id = ?", (google_id,)
    ).fetchone()
    if user:
        session.clear()
        session["username"] = user["username"]
        session["user_id"]  = user["id"]
        conn.close()
        return redirect(url_for("home"))

    # Email already registered via password -- link Google to it
    if email:
        user = conn.execute(
            "SELECT * FROM users WHERE email = ?", (email,)
        ).fetchone()
        if user:
            conn.execute(
                "UPDATE users SET google_id = ? WHERE id = ?",
                (google_id, user["id"]),
            )
            conn.commit()
            session.clear()
            session["username"] = user["username"]
            session["user_id"]  = user["id"]
            conn.close()
            return redirect(url_for("home"))

    # New user -- generate a username from their Google display name
    base = re.sub(r"[^a-zA-Z0-9_]", "", name.replace(" ", "_"))[:17] or "user"
    username = base
    counter = 1
    while conn.execute("SELECT 1 FROM users WHERE username = ?", (username,)).fetchone():
        username = f"{base}_{counter}"
        counter += 1

    try:
        conn.execute(
            "INSERT INTO users (username, email, google_id) VALUES (?, ?, ?)",
            (username, email or None, google_id),
        )
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        flash("Could not create your account. Please try again.")
        return redirect(url_for("signup"))

    user = conn.execute(
        "SELECT * FROM users WHERE google_id = ?", (google_id,)
    ).fetchone()
    session.clear()
    session["username"] = user["username"]
    session["user_id"]  = user["id"]
    conn.close()

    flash(f"Welcome! Your username is: {username}")
    return redirect(url_for("home"))


# "€"€ Social: Solutions (buddy system + chat) + Profile "€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€

def _buddy_status(conn, me, other):
    """Return 'accepted','pending_sent','pending_recv','none'."""
    row = conn.execute(
        "SELECT requester, status FROM buddies WHERE (requester=? AND recipient=?) OR (requester=? AND recipient=?)",
        (me, other, other, me)
    ).fetchone()
    if not row: return "none"
    if row["status"] == "accepted": return "accepted"
    if row["requester"] == me: return "pending_sent"
    return "pending_recv"

def _display_name(row):
    try:    return b64_decode(row["full_name"]) if row["full_name"] else row["username"]
    except: return row["username"]


# 
# CHALLENGES
# 

# Challenges page moved to mobile app (toogoodapp)


@app.route("/perfect/api/challenges")
def api_challenges():
    me = _api_user()
    if not me:
        return jsonify({"ok": False}), 401
    conn = get_db_connection()
    pk   = datetime.utcnow().strftime("%Y-%m")
    rows = conn.execute("SELECT * FROM challenges WHERE period_key=? ORDER BY id", (pk,)).fetchall()

    challenges = []
    for ch in rows:
        ch = dict(ch)
        my_prog = _challenge_progress(conn, me, ch)
        pct     = min(100, round(my_prog / ch["target"] * 100))

        # Top 5 leaderboard
        all_users = conn.execute("SELECT username FROM users").fetchall()
        board = []
        for u in all_users:
            prog = _challenge_progress(conn, u["username"], ch)
            if prog > 0:
                board.append({"username": u["username"], "progress": prog})
        board.sort(key=lambda x: -x["progress"])
        for i, b in enumerate(board):
            b["rank"] = i + 1
            b["is_me"] = b["username"] == me
        my_rank = next((b["rank"] for b in board if b["is_me"]), len(board) + 1)

        challenges.append({
            "id":          ch["id"],
            "title":       ch["title"],
            "description": ch["description"],
            "metric":      ch["metric"],
            "target":      ch["target"],
            "my_progress": my_prog,
            "pct":         pct,
            "my_rank":     my_rank,
            "top5":        board[:5],
            "completed":   my_prog >= ch["target"],
        })
    conn.close()
    return jsonify({"ok": True, "challenges": challenges})


@app.route("/perfect/api/challenges/<int:cid>/leaderboard")
def api_challenge_leaderboard(cid):
    me = _api_user()
    if not me:
        return jsonify({"ok": False}), 401
    conn = get_db_connection()
    ch   = conn.execute("SELECT * FROM challenges WHERE id=?", (cid,)).fetchone()
    if not ch:
        conn.close()
        return jsonify({"ok": False, "error": "not found"}), 404
    ch = dict(ch)

    all_users = conn.execute("SELECT username, full_name FROM users").fetchall()
    board = []
    for u in all_users:
        prog = _challenge_progress(conn, u["username"], ch)
        try:    name = b64_decode(u["full_name"]) if u["full_name"] else u["username"]
        except: name = u["username"]
        board.append({"username": u["username"], "name": name, "progress": prog, "target": ch["target"]})
    board.sort(key=lambda x: -x["progress"])
    for i, b in enumerate(board):
        b["rank"]  = i + 1
        b["is_me"] = b["username"] == me
        b["pct"]   = min(100, round(b["progress"] / ch["target"] * 100))
    conn.close()
    return jsonify({"ok": True, "board": board, "challenge": ch})


# 
# CLUBS
# 

# Clubs page moved to mobile app (toogoodapp)


@app.route("/perfect/api/clubs")
def api_clubs():
    me = _api_user()
    if not me:
        return jsonify({"ok": False}), 401
    conn = get_db_connection()

    my_rows = conn.execute("""
        SELECT c.id, c.name, c.description, c.creator,
               (SELECT COUNT(*) FROM club_members WHERE club_id=c.id) AS member_count
        FROM clubs c
        JOIN club_members cm ON cm.club_id=c.id AND cm.username=?
        ORDER BY c.created_at DESC
    """, (me,)).fetchall()

    discover = conn.execute("""
        SELECT c.id, c.name, c.description, c.creator,
               (SELECT COUNT(*) FROM club_members WHERE club_id=c.id) AS member_count
        FROM clubs c
        WHERE c.id NOT IN (SELECT club_id FROM club_members WHERE username=?)
        ORDER BY member_count DESC, c.created_at DESC
        LIMIT 20
    """, (me,)).fetchall()

    conn.close()
    def fmt(rows):
        return [{"id":r["id"],"name":r["name"],"description":r["description"],
                 "creator":r["creator"],"member_count":r["member_count"]} for r in rows]
    return jsonify({"ok": True, "my_clubs": fmt(my_rows), "discover": fmt(discover)})


@app.route("/perfect/api/clubs/create", methods=["POST"])
def api_clubs_create():
    me = _api_user()
    if not me:
        return jsonify({"ok": False}), 401
    data = request.json or {}
    name = (data.get("name") or "").strip()[:100]
    desc = (data.get("description") or "").strip()[:500]
    if not name:
        return jsonify({"ok": False, "error": "Club name required"})
    conn = get_db_connection()
    cur  = conn.execute("INSERT INTO clubs (name,description,creator) VALUES (?,?,?)", (name, desc, me))
    club_id = cur.lastrowid
    conn.execute("INSERT INTO club_members (club_id,username,role) VALUES (?,?,?)", (club_id, me, "admin"))
    _check_achievements(conn, me)
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "club_id": club_id})


@app.route("/perfect/api/clubs/<int:club_id>/join", methods=["POST"])
def api_clubs_join(club_id):
    me = _api_user()
    if not me:
        return jsonify({"ok": False}), 401
    conn = get_db_connection()
    if not conn.execute("SELECT 1 FROM clubs WHERE id=?", (club_id,)).fetchone():
        conn.close()
        return jsonify({"ok": False, "error": "Club not found"}), 404
    conn.execute("INSERT OR IGNORE INTO club_members (club_id,username) VALUES (?,?)", (club_id, me))
    _check_achievements(conn, me)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/perfect/api/clubs/<int:club_id>/leave", methods=["POST"])
def api_clubs_leave(club_id):
    me = _api_user()
    if not me:
        return jsonify({"ok": False}), 401
    conn = get_db_connection()
    conn.execute("DELETE FROM club_members WHERE club_id=? AND username=?", (club_id, me))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/perfect/api/clubs/<int:club_id>")
def api_clubs_detail(club_id):
    me = _api_user()
    if not me:
        return jsonify({"ok": False}), 401
    conn = get_db_connection()
    club = conn.execute("SELECT * FROM clubs WHERE id=?", (club_id,)).fetchone()
    if not club:
        conn.close()
        return jsonify({"ok": False, "error": "not found"}), 404

    is_member = bool(conn.execute("SELECT 1 FROM club_members WHERE club_id=? AND username=?", (club_id, me)).fetchone())
    members_raw = conn.execute("""
        SELECT cm.username, cm.role, u.full_name,
               COALESCE(u.score_xp,0) AS xp,
               COALESCE(u.score_streak,0) AS streak
        FROM club_members cm JOIN users u ON u.username=cm.username
        WHERE cm.club_id=? ORDER BY xp DESC
    """, (club_id,)).fetchall()

    pk = datetime.utcnow().strftime("%Y-%m")
    members = []
    for m in members_raw:
        try:    name = b64_decode(m["full_name"]) if m["full_name"] else m["username"]
        except: name = m["username"]
        month_xp = conn.execute(
            "SELECT COALESCE(SUM(xp_awarded),0) FROM score_events WHERE username=? AND strftime('%Y-%m',awarded_at)=?",
            (m["username"], pk)
        ).fetchone()[0]
        members.append({"username":m["username"],"name":name,"role":m["role"],
                         "xp":m["xp"],"streak":m["streak"],"month_xp":int(month_xp),
                         "is_me":m["username"]==me})
    members.sort(key=lambda x: -x["month_xp"])
    for i, m in enumerate(members): m["rank"] = i + 1

    conn.close()
    return jsonify({"ok": True, "club": dict(club), "members": members, "is_member": is_member})


@app.route("/perfect/api/clubs/<int:club_id>/feed")
def api_clubs_feed(club_id):
    me = _api_user()
    if not me:
        return jsonify({"ok": False}), 401
    conn = get_db_connection()
    if not conn.execute("SELECT 1 FROM club_members WHERE club_id=? AND username=?", (club_id, me)).fetchone():
        conn.close()
        return jsonify({"ok": False, "error": "Not a member"}), 403

    members = [r["username"] for r in conn.execute(
        "SELECT username FROM club_members WHERE club_id=?", (club_id,)
    ).fetchall()]
    placeholders = ",".join("?" * len(members))
    events = conn.execute(f"""
        SELECT se.id, se.username, se.event_type, se.xp_awarded, se.awarded_at, se.note,
               u.full_name
        FROM score_events se JOIN users u ON u.username=se.username
        WHERE se.username IN ({placeholders})
        ORDER BY se.awarded_at DESC LIMIT 60
    """, members).fetchall()

    feed = []
    for e in events:
        try:    name = b64_decode(e["full_name"]) if e["full_name"] else e["username"]
        except: name = e["username"]
        feed.append({"id":e["id"],"username":e["username"],"name":name,
                     "event_type":e["event_type"],"xp":round(e["xp_awarded"]),"at":e["awarded_at"],"note":e["note"]})
    conn.close()
    return jsonify({"ok": True, "feed": feed})


# 
# ACHIEVEMENTS
# 

@app.route("/perfect/api/achievements")
def api_achievements():
    me = _api_user()
    if not me:
        return jsonify({"ok": False}), 401
    conn = get_db_connection()
    rows = conn.execute("SELECT badge, earned_at FROM achievements WHERE username=? ORDER BY earned_at", (me,)).fetchall()
    conn.close()
    result = []
    for r in rows:
        meta = ACHIEVEMENT_META.get(r["badge"], {"emoji":"","title":r["badge"],"desc":""})
        result.append({"badge":r["badge"],"earned_at":r["earned_at"][:10], **meta})
    return jsonify({"ok": True, "achievements": result})


# 
# SEGMENTS
# 

SEGMENT_CATEGORIES = {
    "run":    {"emoji": "", "label": "Running"},
    "gym":    {"emoji": "'", "label": "Gym"},
    "cycle":  {"emoji": "", "label": "Cycling"},
    "swim":   {"emoji": "", "label": "Swimming"},
    "general":{"emoji": "", "label": "General"},
}
SEGMENT_METRICS = {
    "time":     {"label": "Time (minutes)", "unit": "min",  "best": "low"},
    "reps":     {"label": "Reps / Count",   "unit": "reps", "best": "high"},
    "distance": {"label": "Distance (km)",  "unit": "km",   "best": "high"},
}

def _fmt_value(value, metric):
    if metric == "time":
        mins = int(value)
        secs = int(round((value - mins) * 60))
        return f"{mins}m {secs:02d}s" if secs else f"{mins} min"
    elif metric == "reps":
        return f"{int(value)} reps"
    else:
        return f"{value:.2f} km"

def _segment_leaderboard(conn, segment_id, metric):
    """Return personal-best leaderboard for a segment."""
    agg = "MIN(value)" if SEGMENT_METRICS[metric]["best"] == "low" else "MAX(value)"
    rows = conn.execute(f"""
        SELECT se.username, {agg} AS pb, MAX(se.logged_at) AS last_at, u.full_name
        FROM segment_efforts se JOIN users u ON u.username=se.username
        WHERE se.segment_id=?
        GROUP BY se.username
        ORDER BY pb {'ASC' if SEGMENT_METRICS[metric]['best']=='low' else 'DESC'}
    """, (segment_id,)).fetchall()
    board = []
    for i, r in enumerate(rows):
        try:    name = b64_decode(r["full_name"]) if r["full_name"] else r["username"]
        except: name = r["username"]
        board.append({
            "rank":     i + 1,
            "username": r["username"],
            "name":     name,
            "pb":       r["pb"],
            "pb_fmt":   _fmt_value(r["pb"], metric),
            "last_at":  (r["last_at"] or "")[:10],
        })
    return board


# Segments page moved to mobile app (toogoodapp)


@app.route("/perfect/api/segments")
def api_segments_list():
    me = _api_user()
    if not me:
        return jsonify({"ok": False}), 401
    conn = get_db_connection()
    rows = conn.execute("""
        SELECT s.*,
               (SELECT COUNT(DISTINCT username) FROM segment_efforts WHERE segment_id=s.id) AS athlete_count,
               (SELECT COUNT(*) FROM segment_efforts WHERE segment_id=s.id) AS effort_count,
               (SELECT COUNT(*) FROM segment_efforts WHERE segment_id=s.id AND username=?) AS my_efforts
        FROM segments s ORDER BY effort_count DESC, s.created_at DESC
    """, (me,)).fetchall()
    conn.close()
    result = []
    for r in rows:
        cat  = SEGMENT_CATEGORIES.get(r["category"], SEGMENT_CATEGORIES["general"])
        metr = SEGMENT_METRICS.get(r["metric"], SEGMENT_METRICS["time"])
        result.append({
            "id":           r["id"],
            "name":         r["name"],
            "description":  r["description"],
            "category":     r["category"],
            "cat_emoji":    cat["emoji"],
            "cat_label":    cat["label"],
            "metric":       r["metric"],
            "metric_unit":  metr["unit"],
            "creator":      r["creator"],
            "athlete_count":r["athlete_count"],
            "effort_count": r["effort_count"],
            "my_efforts":   r["my_efforts"],
            "is_mine":      r["creator"] == me,
        })
    return jsonify({"ok": True, "segments": result})


@app.route("/perfect/api/segments/create", methods=["POST"])
def api_segments_create():
    me = _api_user()
    if not me:
        return jsonify({"ok": False}), 401
    data = request.json or {}
    name     = (data.get("name") or "").strip()[:100]
    desc     = (data.get("description") or "").strip()[:500]
    category = data.get("category", "general")
    metric   = data.get("metric", "time")
    if not name:
        return jsonify({"ok": False, "error": "Segment name required"})
    if category not in SEGMENT_CATEGORIES:
        category = "general"
    if metric not in SEGMENT_METRICS:
        metric = "time"
    conn = get_db_connection()
    cur  = conn.execute(
        "INSERT INTO segments (name,description,category,metric,creator) VALUES (?,?,?,?,?)",
        (name, desc, category, metric, me)
    )
    seg_id = cur.lastrowid
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "segment_id": seg_id})


@app.route("/perfect/api/segments/<int:seg_id>")
def api_segment_detail(seg_id):
    me = _api_user()
    if not me:
        return jsonify({"ok": False}), 401
    conn = get_db_connection()
    seg  = conn.execute("SELECT * FROM segments WHERE id=?", (seg_id,)).fetchone()
    if not seg:
        conn.close()
        return jsonify({"ok": False, "error": "not found"}), 404
    seg    = dict(seg)
    metric = seg["metric"]
    board  = _segment_leaderboard(conn, seg_id, metric)

    my_efforts = conn.execute(
        "SELECT value, note, logged_at FROM segment_efforts WHERE segment_id=? AND username=? ORDER BY logged_at DESC LIMIT 10",
        (seg_id, me)
    ).fetchall()
    my_history = [{"value": r["value"], "fmt": _fmt_value(r["value"], metric),
                   "note": r["note"], "at": (r["logged_at"] or "")[:10]} for r in my_efforts]

    my_pb = next((b for b in board if b["username"] == me), None)
    kom   = board[0] if board else None

    cat  = SEGMENT_CATEGORIES.get(seg["category"], SEGMENT_CATEGORIES["general"])
    metr = SEGMENT_METRICS.get(metric, SEGMENT_METRICS["time"])
    conn.close()
    return jsonify({
        "ok":       True,
        "segment":  {**seg, "cat_emoji": cat["emoji"], "cat_label": cat["label"],
                     "metric_label": metr["label"], "metric_unit": metr["unit"],
                     "metric_best": metr["best"]},
        "board":    board,
        "kom":      kom,
        "my_pb":    my_pb,
        "my_history": my_history,
        "is_mine":  seg["creator"] == me,
    })


@app.route("/perfect/api/segments/<int:seg_id>/effort", methods=["POST"])
def api_segment_effort(seg_id):
    me = _api_user()
    if not me:
        return jsonify({"ok": False}), 401
    data = request.json or {}
    try:
        value = float(data.get("value", 0))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "Invalid value"})
    if value <= 0:
        return jsonify({"ok": False, "error": "Value must be greater than 0"})
    note = (data.get("note") or "").strip() or None
    conn = get_db_connection()
    seg  = conn.execute("SELECT metric FROM segments WHERE id=?", (seg_id,)).fetchone()
    if not seg:
        conn.close()
        return jsonify({"ok": False, "error": "Segment not found"}), 404
    conn.execute(
        "INSERT INTO segment_efforts (segment_id,username,value,note) VALUES (?,?,?,?)",
        (seg_id, me, value, note)
    )
    conn.commit()
    metric = seg["metric"]
    board  = _segment_leaderboard(conn, seg_id, metric)
    my_pb  = next((b for b in board if b["username"] == me), None)
    kom    = board[0] if board else None
    conn.close()
    is_kom = kom and kom["username"] == me
    is_pr  = my_pb and my_pb["pb"] == value if SEGMENT_METRICS[metric]["best"] == "high" else (my_pb and my_pb["pb"] == value)
    return jsonify({
        "ok":      True,
        "fmt":     _fmt_value(value, metric),
        "rank":    my_pb["rank"] if my_pb else None,
        "is_kom":  is_kom,
        "kom_name": kom["name"] if kom else None,
    })


@app.route("/perfect/api/segments/<int:seg_id>/delete", methods=["POST"])
def api_segment_delete(seg_id):
    me = _api_user()
    if not me:
        return jsonify({"ok": False}), 401
    conn = get_db_connection()
    seg  = conn.execute("SELECT creator FROM segments WHERE id=?", (seg_id,)).fetchone()
    if not seg or seg["creator"] != me:
        conn.close()
        return jsonify({"ok": False, "error": "Not authorised"}), 403
    conn.execute("DELETE FROM segment_efforts WHERE segment_id=?", (seg_id,))
    conn.execute("DELETE FROM segments WHERE id=?", (seg_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/perfect/socials")
def perfect_socials():
    if not session.get("username"):
        return redirect(url_for("login"))
    return render_template("perfect/socials.html",
        username=session["username"],
        full_name=session.get("full_name", ""),
        email=session.get("email", ""),
        created_at=session.get("created_at", ""),
        page_active="socials")


@app.route("/perfect/profile")
@app.route("/perfect/profile/<target>")
def perfect_profile(target=None):
    if not session.get("username"):
        return redirect(url_for("login"))
    me = session["username"]
    if not target or target == me:
        target = me
    conn = get_db_connection()
    row = conn.execute(
        "SELECT username, full_name, created_at, goal, score_xp, score_streak FROM users WHERE username=?",
        (target,)
    ).fetchone()
    if not row:
        conn.close()
        return "User not found", 404
    buddy_count = conn.execute(
        "SELECT COUNT(*) FROM buddies WHERE (requester=? OR recipient=?) AND status='accepted'",
        (target, target)
    ).fetchone()[0]
    status = _buddy_status(conn, me, target) if target != me else "self"
    conn.close()
    lvl, _, _ = _compute_level(row["score_xp"] or 0)
    goal_labels = {
        "weight_loss": "Weight Loss", "muscle_gain": "Muscle Gain",
        "maintenance": "Maintenance", "fat_loss": "Fat Loss",
        "weight_gain": "Weight Gain", "general_fitness": "General Fitness",
    }
    return render_template("perfect/profile.html",
        username=session["username"],
        full_name=session.get("full_name", ""),
        email=session.get("email", ""),
        created_at=session.get("created_at", ""),
        page_active="profile",
        target_username=target,
        target_name=_display_name(row),
        target_joined=str(row["created_at"] or "")[:10],
        target_goal=goal_labels.get(row["goal"] or "", row["goal"] or "--"),
        target_level=lvl,
        target_xp=round(row["score_xp"] or 0),
        target_streak=row["score_streak"] or 0,
        buddy_count=buddy_count,
        buddy_status=status,
        is_me=(target == me))


# "€"€ Social API "€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€"€

@app.route("/perfect/api/users/search")
def api_users_search():
    me = _api_user()
    if not me: return jsonify({"error":"Not logged in"}), 401
    q  = request.args.get("q", "").strip().lower()
    if len(q) < 2: return jsonify({"results": []})
    conn = get_db_connection()
    rows = conn.execute(
        "SELECT username, full_name, score_xp FROM users WHERE username != ? AND LOWER(username) LIKE ? LIMIT 12",
        (me, f"%{q}%")
    ).fetchall()
    results = []
    for r in rows:
        lvl, _, _ = _compute_level(r["score_xp"] or 0)
        results.append({
            "username": r["username"],
            "name": _display_name(r),
            "level": lvl,
            "status": _buddy_status(conn, me, r["username"]),
        })
    conn.close()
    return jsonify({"results": results})


@app.route("/perfect/api/buddy/request", methods=["POST"])
def api_buddy_request():
    me = _api_user()
    if not me: return jsonify({"error":"Not logged in"}), 401
    other = (request.get_json(force=True) or {}).get("username", "").strip()
    if not other or other == me: return jsonify({"error":"Invalid user"}), 400
    conn = get_db_connection()
    exists = conn.execute("SELECT id FROM users WHERE username=?", (other,)).fetchone()
    if not exists:
        conn.close(); return jsonify({"error":"User not found"}), 404
    try:
        conn.execute("INSERT INTO buddies (requester, recipient) VALUES (?,?)", (me, other))
        conn.commit()
    except Exception:
        conn.close(); return jsonify({"error":"Request already exists"}), 409
    conn.close()
    return jsonify({"ok": True})


@app.route("/perfect/api/buddy/respond", methods=["POST"])
def api_buddy_respond():
    me = _api_user()
    if not me: return jsonify({"error":"Not logged in"}), 401
    data = request.get_json(force=True) or {}
    other  = data.get("username", "").strip()
    action = data.get("action", "")   # "accept" | "reject"
    if not other or action not in ("accept", "reject"):
        return jsonify({"error":"Bad request"}), 400
    conn = get_db_connection()
    if action == "accept":
        conn.execute("UPDATE buddies SET status='accepted' WHERE requester=? AND recipient=? AND status='pending'",
                     (other, me))
    else:
        conn.execute("DELETE FROM buddies WHERE requester=? AND recipient=? AND status='pending'", (other, me))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/perfect/api/buddy/remove", methods=["POST"])
def api_buddy_remove():
    me = _api_user()
    if not me: return jsonify({"error":"Not logged in"}), 401
    other = (request.get_json(force=True) or {}).get("username", "").strip()
    conn  = get_db_connection()
    conn.execute("DELETE FROM buddies WHERE (requester=? AND recipient=?) OR (requester=? AND recipient=?)",
                 (me, other, other, me))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/perfect/api/buddies")
def api_buddies():
    me = _api_user()
    if not me: return jsonify({"error":"Not logged in"}), 401
    conn = get_db_connection()
    # accepted buddies
    rows = conn.execute("""
        SELECT u.username, u.full_name, u.score_xp, u.score_streak
        FROM buddies b
        JOIN users u ON (b.requester=u.username OR b.recipient=u.username) AND u.username != ?
        WHERE (b.requester=? OR b.recipient=?) AND b.status='accepted'
    """, (me, me, me)).fetchall()
    buddies = []
    for r in rows:
        lvl, _, _ = _compute_level(r["score_xp"] or 0)
        # latest activity timestamp
        last = conn.execute(
            "SELECT awarded_at FROM score_events WHERE username=? ORDER BY awarded_at DESC LIMIT 1",
            (r["username"],)
        ).fetchone()
        buddies.append({"username": r["username"], "name": _display_name(r),
                        "level": lvl, "streak": r["score_streak"] or 0,
                        "last_active": str(last["awarded_at"])[:16] if last else None})
    # pending incoming requests
    inc = conn.execute("""
        SELECT u.username, u.full_name, u.score_xp, b.created_at FROM buddies b
        JOIN users u ON b.requester=u.username
        WHERE b.recipient=? AND b.status='pending'
        ORDER BY b.created_at DESC
    """, (me,)).fetchall()
    incoming = [{"username": r["username"], "name": _display_name(r),
                 "level": _compute_level(r["score_xp"] or 0)[0]} for r in inc]
    # pending outgoing
    out = conn.execute("""
        SELECT u.username, u.full_name FROM buddies b
        JOIN users u ON b.recipient=u.username
        WHERE b.requester=? AND b.status='pending'
    """, (me,)).fetchall()
    outgoing = [{"username": r["username"], "name": _display_name(r)} for r in out]
    conn.close()
    return jsonify({"ok": True, "buddies": buddies, "incoming": incoming, "outgoing": outgoing})


@app.route("/perfect/api/activity/<other>")
def api_activity(other):
    """Return a buddy's recent activity feed - only accessible to accepted buddies."""
    me = _api_user()
    if not me: return jsonify({"error":"Not logged in"}), 401
    conn = get_db_connection()
    if _buddy_status(conn, me, other) != "accepted":
        conn.close(); return jsonify({"error":"Not buddies"}), 403

    # Basic profile info
    row = conn.execute(
        "SELECT username, full_name, score_xp, score_streak, goal, exercise_types, "
        "exercise_days_per_week, fitness_level FROM users WHERE username=?", (other,)
    ).fetchone()
    if not row: conn.close(); return jsonify({"error":"Not found"}), 404

    lvl, xp_in, xp_need = _compute_level(row["score_xp"] or 0)

    # Recent score events (activity)
    events = conn.execute(
        "SELECT event_type, xp_awarded, awarded_at, note FROM score_events "
        "WHERE username=? ORDER BY awarded_at DESC LIMIT 40", (other,)
    ).fetchall()

    conn.close()

    goal_labels = {
        "weight_loss":"Weight Loss","muscle_gain":"Muscle Gain","maintenance":"Maintenance",
        "fat_loss":"Fat Loss","weight_gain":"Weight Gain","general_fitness":"General Fitness",
    }

    try:    ex_types = row["exercise_types"] or ""
    except: ex_types = ""

    events_raw = conn.execute(
        "SELECT id, event_type, xp_awarded, awarded_at, note FROM score_events "
        "WHERE username=? ORDER BY awarded_at DESC LIMIT 40", (other,)
    ).fetchall()

    feed = []
    icons  = {"login":"","food_log":"","calorie_goal":"","exercise":"","blood_report":"'"}
    labels = {"login":"Checked in","food_log":"Logged their meals","calorie_goal":"Hit their calorie goal",
              "exercise":"Completed a workout","blood_report":"Analyzed a blood report"}
    for e in events_raw:
        eid = e["id"]
        reacts = conn.execute(
            "SELECT emoji, COUNT(*) as n FROM activity_reactions WHERE event_id=? GROUP BY emoji", (eid,)
        ).fetchall()
        my_react = conn.execute(
            "SELECT emoji FROM activity_reactions WHERE event_id=? AND reactor=?", (eid, me)
        ).fetchone()
        comments = conn.execute(
            "SELECT commenter, body, created_at FROM activity_comments WHERE event_id=? ORDER BY created_at ASC", (eid,)
        ).fetchall()
        feed.append({
            "id":      eid,
            "type":    e["event_type"],
            "icon":    icons.get(e["event_type"], ""),
            "label":   labels.get(e["event_type"], e["event_type"]),
            "xp":      round(e["xp_awarded"]),
            "at":      str(e["awarded_at"])[:16],
            "note":    e["note"],
            "reactions": [{"emoji": r["emoji"], "count": r["n"]} for r in reacts],
            "my_react":  my_react["emoji"] if my_react else None,
            "comments":  [{"who": c["commenter"], "body": c["body"], "at": str(c["created_at"])[:16]} for c in comments],
        })

    conn.close()

    return jsonify({
        "ok": True,
        "name":     _display_name(row),
        "level":    lvl,
        "xp":       round(row["score_xp"] or 0),
        "streak":   row["score_streak"] or 0,
        "goal":     goal_labels.get(row["goal"] or "", row["goal"] or "--"),
        "exercise_types": ex_types,
        "fitness_level":  row["fitness_level"] or "--",
        "feed":     feed,
    })


@app.route("/perfect/api/activity/react", methods=["POST"])
def api_react():
    me = _api_user()
    if not me: return jsonify({"error":"Not logged in"}), 401
    data = request.get_json(force=True) or {}
    event_id = data.get("event_id")
    emoji    = data.get("emoji", "'")
    if not event_id: return jsonify({"error":"Missing event_id"}), 400
    # verify buddy access
    conn = get_db_connection()
    owner = conn.execute("SELECT username FROM score_events WHERE id=?", (event_id,)).fetchone()
    if not owner: conn.close(); return jsonify({"error":"Not found"}), 404
    if owner["username"] != me and _buddy_status(conn, me, owner["username"]) != "accepted":
        conn.close(); return jsonify({"error":"Not buddies"}), 403
    # toggle: if same emoji exists, remove; else upsert
    existing = conn.execute("SELECT emoji FROM activity_reactions WHERE event_id=? AND reactor=?", (event_id, me)).fetchone()
    if existing and existing["emoji"] == emoji:
        conn.execute("DELETE FROM activity_reactions WHERE event_id=? AND reactor=?", (event_id, me))
        reacted = False
    else:
        conn.execute("INSERT OR REPLACE INTO activity_reactions (event_id, reactor, emoji) VALUES (?,?,?)", (event_id, me, emoji))
        reacted = True
    conn.commit()
    count = conn.execute("SELECT COUNT(*) FROM activity_reactions WHERE event_id=? AND emoji=?", (event_id, emoji)).fetchone()[0]
    conn.close()
    return jsonify({"ok": True, "reacted": reacted, "count": count})


@app.route("/perfect/api/activity/comment", methods=["POST"])
def api_comment():
    me = _api_user()
    if not me: return jsonify({"error":"Not logged in"}), 401
    data = request.get_json(force=True) or {}
    event_id = data.get("event_id")
    body     = (data.get("body") or "").strip()
    if not event_id or not body: return jsonify({"error":"Missing fields"}), 400
    if len(body) > 200: return jsonify({"error":"Too long"}), 400
    conn = get_db_connection()
    owner = conn.execute("SELECT username FROM score_events WHERE id=?", (event_id,)).fetchone()
    if not owner: conn.close(); return jsonify({"error":"Not found"}), 404
    if owner["username"] != me and _buddy_status(conn, me, owner["username"]) != "accepted":
        conn.close(); return jsonify({"error":"Not buddies"}), 403
    conn.execute("INSERT INTO activity_comments (event_id, commenter, body) VALUES (?,?,?)", (event_id, me, body))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


if __name__ == "__main__":
    _debug = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(host="0.0.0.0", port=5000, debug=_debug)


